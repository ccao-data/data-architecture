# Export dbt models to Excel files ("reports").
#
# Run `python scripts/qc_report.py --help` for details.

import argparse
import contextlib
import io
import json
import os
import pathlib
import shutil

import pandas as pd
import pyathena
from dbt.cli.main import dbtRunner
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

DBT = dbtRunner()
CLI_DESCRIPTION = """Export dbt models to Excel files ("reports").

A few configuration values can be set on any model used to generate reports:

    * config.meta.report_name (optional): The name of the report to generate. This will be used to set the filename for the output report. File
      extensions are not necessary since all reports are output as .xlsx files. If unset, defaults to the name of the model.

    * config.meta.template (optional): The filename an Excel template to use when generating the report. Templates should be stored in the
      reports/templates/ directory and should include header rows. If unset, will search for a template with the same name as the model; if no
      template is found, defaults to a simple layout with filterable columns and striped rows.
"""  # noqa: E501
CLI_EXAMPLE = """Example usage to output the qc_report_town_close report for Hyde Park township:

    python scripts/qc_report.py --select tag:qc_report_town_close --vars '{qc_report_town_code: 70}' --rebuild
"""  # noqa: E501


def main():
    parser = argparse.ArgumentParser(
        description=CLI_DESCRIPTION,
        epilog=CLI_EXAMPLE,
        # Parse the description and epilog as raw text so that newlines
        # get preserved
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument(
        "--select",
        required=True,
        nargs="*",
        help="One or more dbt select statements to use for filtering reports",
    )
    parser.add_argument(
        "--target",
        required=False,
        default="dev",
        help="dbt target to use for querying data, defaults to 'dev'",
    )
    parser.add_argument(
        "--rebuild",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Rebuild models before exporting reports",
    )
    parser.add_argument(
        "--vars",
        required=False,
        help=(
            "Variable overrides to pass to `dbt build`. Expects a YAML "
            "string. Useful for e.g. overriding the township that is used "
            "to generate a report. Requires that --rebuild be set"
        ),
    )

    args = parser.parse_args()
    target = args.target
    select = args.select
    rebuild = args.rebuild
    vars = args.vars

    if vars and not rebuild:
        raise ValueError("--rebuild must be set if --vars is set")

    if rebuild:
        print("Rebuilding models")
        dbt_run_args = [
            "run",
            "--target",
            target,
            "--select",
            *select,
        ]
        if vars:
            dbt_run_args += ["--vars", vars]

        dbt_run_result = DBT.invoke(dbt_run_args)
        if not dbt_run_result.success:
            print("Encountered error in `dbt run` call")
            raise ValueError(dbt_run_result.exception)

    print("Listing models to select for export")
    dbt_output = io.StringIO()
    with contextlib.redirect_stdout(dbt_output):
        dbt_list_result = DBT.invoke(
            [
                "--quiet",
                "list",
                "--target",
                target,
                "--resource-types",
                "model",
                "--output",
                "json",
                "--output-keys",
                "name",
                "config",
                "relation_name",
                "--select",
                *select,
            ]
        )

    if not dbt_list_result.success:
        print("Encountered error in `dbt list` call")
        raise ValueError(dbt_list_result.exception)

    # Output is formatted as a list of newline-separated JON objects
    models = [
        json.loads(model_dict_str)
        for model_dict_str in dbt_output.getvalue().split("\n")
        # Filter out empty strings caused by trailing newlines
        if model_dict_str
    ]

    if not models:
        raise ValueError(f"No models found for the --select value '{select}'")

    print(
        "The following models will be exported to reports: "
        f"{', '.join(model['name'] for model in models)}"
    )

    conn = pyathena.connect(
        s3_staging_dir=os.getenv(
            "AWS_ATHENA_S3_STAGING_DIR",
            "s3://ccao-dbt-athena-results-us-east-1",
        ),
        region_name=os.getenv("AWS_ATHENA_REGION_NAME", "us-east-1"),
    )

    for model in models:
        # Extract useful model metadata from the columns we queried in
        # the `dbt list` call above
        model_name = model["name"]
        relation_name = model["relation_name"]
        report_name = model["config"]["meta"].get("report_name") or model_name
        template = (
            model["config"]["meta"].get("template") or f"{model_name}.xlsx"
        )

        # Define inputs and outputs for report based on model metadata
        template_path = os.path.join("reports", "templates", template)
        template_exists = os.path.isfile(template_path)
        output_path = os.path.join("reports", "output", f"{report_name}.xlsx")

        print(f"Querying data for model {model_name}")
        model_df = pd.read_sql(f"SELECT * FROM {relation_name}", conn)

        # Delete the output file if one already exists
        pathlib.Path(output_path).unlink(missing_ok=True)

        if template_exists:
            print(f"Using template file at {template_path}")
            shutil.copyfile(template_path, output_path)
        else:
            print("No template file exists; creating a workbook from scratch")

        writer_kwargs = (
            {"mode": "a", "if_sheet_exists": "overlay"}
            if template_exists
            else {}
        )
        with pd.ExcelWriter(
            output_path, engine="openpyxl", **writer_kwargs
        ) as writer:
            # TODO: Support sheet name customization
            sheet_name = "Query Results" if template_exists else "Sheet1"
            model_df.to_excel(
                writer,
                sheet_name=sheet_name,
                header=False if template_exists else True,
                index=False,
                startrow=1 if template_exists else 0,
            )

            # Add a table for data filtering
            sheet = writer.sheets[sheet_name]
            table = Table(
                displayName="Query_Results",
                ref=(
                    f"A1:{get_column_letter(sheet.max_column)}"
                    f"{str(sheet.max_row)}"
                ),
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium7", showRowStripes=True
            )
            sheet.add_table(table)

        print(f"Wrote report from model {model_name} to {output_path}")


if __name__ == "__main__":
    main()
