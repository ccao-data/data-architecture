# Shared utilities for exporting models
import contextlib
import dataclasses
import decimal
import io
import json
import os
import pathlib
import shutil
import typing

import pandas as pd
import pyathena
import pyathena.pandas.cursor
from dbt.cli.main import dbtRunner
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Shared object for running dbt CLI commands
DBT = dbtRunner()


def export_models(
    target: str = "dev",
    select: list[str] | None = None,
    selector: str | None = None,
    rebuild: bool = False,
    where: str | None = None,
    output_dir: str | None = None,
) -> list[pathlib.Path]:
    """
    Export a group of models to Excel workbooks in the output directory
    `export/output/`.

    Convenience function that wraps the lower-level `query_models_for_export`
    and `save_model_to_workbook` functions. Use those functions directly if
    you would like to further modify query results or output configurations
    between the query step and the save step.

    Arguments:

        * target (str): dbt target to use for querying model data, defaults to
            "dev"
        * select (list[str]): One or more dbt --select statements to
            use for filtering models. One of --select or --selector is
            required, and if both are set, --selector takes precedence
        * selector (str): A selector name to use for filtering
            models, as defined in selectors.yml. Takes precedence over
            --select if both are set
        * rebuild (bool): Rebuild models before exporting, defaults to False
        * where (str): Optional SQL expression representing a WHERE clause to
            filter models
        * output_dir (str): Optional Unix path to directory where output files
            should be stored

    Returns a list of pathlib.Path objects representing the paths to the files
    that the function created.
    """
    models_for_export = query_models_for_export(
        target=target,
        select=select,
        selector=selector,
        rebuild=rebuild,
        where=where,
    )
    output_paths: list[pathlib.Path] = []
    for model in models_for_export:
        output_path = save_model_to_workbook(model, output_dir)
        output_paths.append(output_path)
    return output_paths


@dataclasses.dataclass
class ModelForExport:
    """Object containing row-level data and table-level metadata for a model
    providing for its export to an Excel workbook."""

    # Model name
    name: str
    # Model dbt tags
    tags: list[str]
    # Dataframe containing model data
    df: pd.DataFrame
    # Optional configs for modifying the display of data in Excel workbook
    export_format: dict
    # Filename for the output file
    export_filename: str
    # Unix-formatted path to a workbook to use as a template
    template_path: str
    # 1-indexed position of the first non-header row
    start_row: int
    # Whether or not to add a data table for sorting and filtering
    add_table: bool


def query_models_for_export(
    target: str = "dev",
    select: list[str] | None = None,
    selector: str | None = None,
    rebuild: bool = False,
    where: str | None = None,
) -> list[ModelForExport]:
    """
    Get data and export metadata for a group of models to export.

    Arguments:

        * target (str): dbt target to use for querying model data, defaults to
            "dev"
        * select (list[str]): One or more dbt --select statements to
            use for filtering models
        * selector (str): A selector name to use for filtering
            models, as defined in selectors.yml. One of `select` or `selector`
            must be set, but they can't both be set
        * rebuild (bool): Rebuild models before exporting, defaults to False
        * where (str): Optional SQL expression representing a WHERE clause to
            filter models

    Returns a list of ModelForExport objects that provide the information
    necessary to save a workbook containing model data.
    """
    if not select and not selector:
        raise ValueError("One of --select or --selector is required")

    select_args = (
        ["--selector", selector] if selector else ["--select", *select]  # type: ignore
    )

    if rebuild:
        dbt_run_args = ["run", "--target", target, *select_args]
        print("Rebuilding models")
        print(f"> dbt {' '.join(dbt_run_args)}")
        dbt_run_result = DBT.invoke(dbt_run_args)
        if not dbt_run_result.success:
            print("Encountered error in `dbt run` call")
            raise ValueError(dbt_run_result.exception)

    print("Listing models to select for export")
    dbt_list_args = [
        "--quiet",
        "list",
        "--target",
        target,
        "--resource-types",
        "model",
        "--output",
        "json",
        "--output-keys",
        "name",
        "config",
        "relation_name",
        "tags",
        *select_args,
    ]
    print(f"> dbt {' '.join(dbt_list_args)}")
    dbt_output = io.StringIO()
    with contextlib.redirect_stdout(dbt_output):
        dbt_list_result = DBT.invoke(dbt_list_args)

    if not dbt_list_result.success:
        print("Encountered error in `dbt list` call")
        raise ValueError(dbt_list_result.exception)

    # Output is formatted as a list of newline-separated JSON objects
    models = [
        json.loads(model_dict_str)
        for model_dict_str in dbt_output.getvalue().split("\n")
        # Filter out empty strings caused by trailing newlines
        if model_dict_str
    ]

    if not models:
        raise ValueError(
            f"No models found for the select option '{' '.join(select_args)}'"
        )

    print(
        "The following models will be exported: "
        f"{', '.join(model['name'] for model in models)}"
    )

    cursor = pyathena.connect(
        s3_staging_dir=os.getenv(
            "AWS_ATHENA_S3_STAGING_DIR",
            "s3://ccao-athena-results-us-east-1/",
        ),
        region_name=os.getenv("AWS_ATHENA_REGION_NAME", "us-east-1"),
        cursor_class=pyathena.pandas.cursor.PandasCursor,
        # Unload query results to speed up execution times
    ).cursor(unload=True)

    models_for_export: list[ModelForExport] = []
    for model in models:
        # Extract useful model metadata from the columns we queried in
        # the `dbt list` call above
        model_name = model["name"]
        tags = model["tags"]
        relation_name = model["relation_name"]
        export_name = model["config"]["meta"].get("export_name") or model_name
        template_config = model["config"]["meta"].get("export_template", {})
        template_name = template_config.get("name") or model_name
        start_row = template_config.get("start_row") or 2
        add_table = template_config.get("add_table") or False

        # Define inputs and outputs for export based on model metadata
        template_path = os.path.join(
            "export", "templates", f"{template_name}.xlsx"
        )

        print(f"Querying data for model {model_name}")
        query = f"SELECT * FROM {relation_name}"
        if where:
            query += f" WHERE {where}"
        print(f"> {query}")
        model_df = cursor.execute(query).as_pandas()

        models_for_export.append(
            ModelForExport(
                name=model_name,
                tags=tags,
                df=model_df,
                export_format=(
                    model["config"]["meta"].get("export_format", {})
                ),
                export_filename=f"{export_name}.xlsx",
                template_path=template_path,
                start_row=start_row,
                add_table=add_table,
            )
        )

    return models_for_export


def save_model_to_workbook(
    model: ModelForExport,
    output_dir: str | None = None,
) -> pathlib.Path:
    """
    Given a ModelForExport object containing row-level data and table-level
    metadata for a model, process that model and save its data to the
    configured location on disk.

    Arguments:

        * model_for_export (ModelForExport): Object containing row-level model
            data and table-level metadata about models that should be saved
        * output_dir (str): Optional Unix path to directory where output files
            should be stored

    Returns a pathlib.Path object representing the path to the file that the
    function created.
    """
    # Compute the full filepath for the output
    output_dirpath = (
        pathlib.Path(output_dir)
        if output_dir
        else pathlib.Path("export/output")
    )
    output_path = output_dirpath / model.export_filename

    # Delete the old version of the output file if one already exists
    output_path.unlink(missing_ok=True)

    template_exists = os.path.isfile(model.template_path)
    if template_exists:
        print(f"Using template file at {model.template_path}")
        shutil.copyfile(model.template_path, output_path)
    else:
        print("No template file exists; creating a workbook from scratch")

    writer_kwargs = (
        {"mode": "a", "if_sheet_exists": "overlay"} if template_exists else {}
    )
    with pd.ExcelWriter(
        output_path, engine="openpyxl", **writer_kwargs
    ) as writer:
        sheet_name = "Sheet1"
        model.df.to_excel(
            writer,
            sheet_name=sheet_name,
            header=False if template_exists else True,
            index=False,
            # Startrow is 0-indexed, whereas the config is 1-indexed
            startrow=model.start_row - 1 if template_exists else 0,
        )
        sheet = writer.sheets[sheet_name]

        # Add a table for data filtering. Only do this if the result set
        # is not empty, because otherwise the empty table will make
        # the Excel workbook invalid
        if model.df.empty:
            print(
                "Skipping formatting for output workbook since result set "
                "is empty"
            )
        else:
            if model.add_table:
                table = Table(
                    displayName="Query_Results",
                    ref=(
                        f"A1:{get_column_letter(sheet.max_column)}"
                        f"{str(sheet.max_row)}"
                    ),
                )
                table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium11", showRowStripes=True
                )
                sheet.add_table(table)

            # Parse column format settings by col index. Since column
            # formatting needs to be applied at the cell level, we'll
            # first parse all format settings for each column, and then
            # we'll iterate every cell once to apply all formatting at the
            # same time
            column_format_by_index = {}

            # If a parid column exists, format it explicitly as a
            # 14-digit number to avoid Excel converting it to scientific
            # notation or stripping out leading zeros
            if "parid" in model.df or "pin" in model.df:
                parid_field = "parid" if "parid" in model.df else "pin"
                parid_index = model.df.columns.get_loc(parid_field)
                column_format_by_index[parid_index] = {
                    "number_format": "00000000000000",
                    # Left align since PINs do not actually need to be
                    # compared by order of magnitude the way that numbers
                    # do
                    "alignment": Alignment(horizontal="left"),
                }

            # Parse any formatting that is configured at the column level.
            # Note that if formatting is configured for a column that we
            # parsed as a parid column above, these settings will override
            # the default parid settings from the block above
            format_config = model.export_format
            if column_configs := format_config.get("columns"):
                for column_config in column_configs:
                    # The column index is required in order to set any
                    # column-level configs
                    col_letter = column_config.get("index")
                    if col_letter is None:
                        raise ValueError(
                            "'index' attribute is required in "
                            "export_format.columns config for "
                            f"model {model.name}"
                        )
                    idx = column_index_from_string(col_letter) - 1
                    # Initialize the config dict for this column if
                    # none exists yet
                    column_format_by_index[idx] = {}
                    # Parse configs if they are present
                    if number_format := column_config.get("number_format"):
                        column_format_by_index[idx]["number_format"] = (
                            number_format
                        )
                    if horiz_align_dir := column_config.get(
                        "horizontal_align"
                    ):
                        column_format_by_index[idx]["alignment"] = Alignment(
                            horizontal=horiz_align_dir
                        )
                    if data_type := column_config.get("data_type"):
                        type_func: typing.Callable
                        if data_type == "int":
                            type_func = int
                        elif data_type == "float":
                            type_func = float
                        elif data_type == "decimal":
                            type_func = decimal.Decimal
                        elif data_type == "str":
                            type_func = str
                        else:
                            raise ValueError(
                                f"data_type '{data_type}' in export_format "
                                f"for model {model.name} is not recognized, "
                                "must be one of 'int', 'float', or 'str'"
                            )
                        column_format_by_index[idx]["data_type"] = type_func

            # Configure the format for blank values in the output workbook. By
            # default, openpyxl will output empty cells as nulls, but we allow
            # users to override this behavior to instead output an empty string
            # using the `format_blanks_as_empty_string` config
            format_blanks_as_empty_string = format_config.get(
                "format_blanks_as_empty_string"
            )
            # openpyxl does not support writing empty strings, so we write a
            # formula that evaluates to an empty string as a workaround. See:
            # https://foss.heptapod.net/openpyxl/openpyxl/-/issues/2174
            blank_value = '=""' if format_blanks_as_empty_string else None

            # Skip header row when applying formatting. We need to
            # catch the special case where there is only one row, or
            # else we will iterate the _cells_ in that row instead of
            # the row when slicing it from 2 : max_row
            non_header_rows = (
                [sheet[model.start_row]]
                if sheet.max_row == model.start_row
                else sheet[model.start_row : sheet.max_row]
            )
            for row in non_header_rows:
                for cell in row:
                    if format_blanks_as_empty_string and (
                        cell.value == "" or cell.value is None
                    ):
                        cell.value = blank_value
                    else:
                        # col_idx is 1-indexed, but our col formats are 0-indexed
                        formats = column_format_by_index.get(
                            cell.col_idx - 1, {}
                        )
                        for attr, val in formats.items():
                            if attr == "data_type":
                                cell.value = (
                                    val(cell.value)
                                    if cell.value != ""
                                    and cell.value is not None
                                    else blank_value
                                )
                            else:
                                setattr(cell, attr, val)

        print(f"Exported model {model.name} to {output_path}")
        return output_path
