#!/usr/bin/env python3
#
# Generates an Excel workbook of dbt test failures that can be shared with
# other teams for review and correction, along with metadata parquet files
# that can be uploaded to S3 for long-term result tracking.
#
# Run `python3 run_iasworld_data_tests.py --help` for detailed
# documentation.

import argparse
import dataclasses
import datetime
import decimal
import enum
import hashlib
import os
import pathlib
import re
import subprocess
import typing

import openpyxl
import openpyxl.cell
import openpyxl.styles
import openpyxl.styles.colors
import openpyxl.utils
import pyarrow as pa
import pyarrow.parquet
import pyathena
import pyathena.arrow.cursor
import pyathena.cursor
import simplejson as json
import yaml
from dbt.artifacts.schemas.results import TestStatus
from dbt.cli.main import dbtRunner
from utils import constants

DBT = dbtRunner()

# Prefix for the URL location of a test in the dbt docs
DOCS_URL_PREFIX = "https://ccao-data.github.io/data-architecture/#!/test"
# The S3 bucket where Athena query results are stored
AWS_ATHENA_S3_STAGING_DIR = os.getenv(
    "AWS_ATHENA_S3_STAGING_DIR", "s3://ccao-athena-results-us-east-1/"
)
# Field names that are used in the output workbook
SOURCE_TABLE_FIELD = "source_table"
DESCRIPTION_FIELD = "description"
TEST_NAME_FIELD = "test_name"
DOCS_URL_FIELD = "docs_url"
TAXYR_FIELD = "taxyr"
PARID_FIELD = "parid"
CARD_FIELD = "card"
LAND_LINE_FIELD = "lline"
TOWNSHIP_FIELD = "township_code"
CLASS_FIELD = "class"
WHO_FIELD = "who"
WEN_FIELD = "wen"
# Mapping that defines category names that should be reported for tests
# based on their generics
TEST_CATEGORIES = {
    "test_accepted_range": "incorrect_values",
    "test_accepted_values": "incorrect_values",
    "test_not_accepted_values": "incorrect_values",
    "test_unique_combination_of_columns": "duplicate_records",
    "test_not_null": "missing_values",
    "test_is_null": "missing_values",
    "test_res_class_matches_pardat": "class_mismatch_or_issue",
}
# Fallback for tests whose category we can't determine from either the
# test name, the `meta.category` attribute, or the TEST_CATEGORIES mapping
DEFAULT_TEST_CATEGORY = "miscellaneous"
# Directory to store failed test caches
TEST_CACHE_DIR = "test_cache"


class Status(enum.Enum):
    """Status of an individual dbt test result."""

    PASS = "pass"
    FAIL = "fail"
    WARN = "warn"

    def __repr__(self) -> str:
        return self.value.upper()


class TestResult:
    """Class to store results for an individual test."""

    def __init__(
        self,
        name: str,
        table_name: str,
        column_name: typing.Optional[str],
        status: Status,
        description: str,
        elapsed_time: decimal.Decimal,
        failing_rows: typing.Optional[typing.List[typing.Dict]] = None,
    ) -> None:
        """
        The failing_rows list should be formatted like the rows
        returned by a csv.DictReader or a DictCursor, i.e. a list of
        dicts mapping `{field_name: row_value}`.
        """
        self.name = name
        self.table_name = table_name
        self.column_name = column_name
        self.status = status
        self.description = description
        self.elapsed_time = elapsed_time
        self.failing_rows: typing.List[typing.Dict] = failing_rows or []

    @property
    def fieldnames(self) -> typing.List[str]:
        """Return a list of strings representing the fieldnames for any
        failing_rows of this test. Returns an empty list if the test
        passed."""
        fieldnames = []
        for row in self.failing_rows:
            for fieldname in row.keys():
                if fieldname not in fieldnames:
                    fieldnames.append(fieldname)
        return fieldnames

    def to_dict(self) -> typing.Dict:
        """Serialize the TestResult object as a dictionary."""
        return {
            "name": self.name,
            "table_name": self.table_name,
            "column_name": self.column_name,
            "status": self.status.value,
            "description": self.description,
            "elapsed_time": self.elapsed_time,
            "failing_rows": self.failing_rows,
        }

    @classmethod
    def from_dict(cls, result_dict: typing.Dict) -> "TestResult":
        """Deserialize a TestResult object from a dictionary."""
        return TestResult(
            name=result_dict["name"],
            table_name=result_dict["table_name"],
            column_name=result_dict["column_name"],
            status=Status(result_dict["status"]),
            description=result_dict["description"],
            elapsed_time=result_dict["elapsed_time"],
            failing_rows=result_dict["failing_rows"],
        )

    def __repr__(self) -> str:
        return f"TestResult({self.to_dict()!r})"

    def split_by_township(self) -> typing.List["TownshipTestResult"]:
        """Split out this TestResult object into one or more
        TownshipTestResults based on the township code of each failing row. If
        there are no failing rows, or if all the failing rows have a null
        township, the return value will be a list of one TownshipTestResult
        whose township_code is None."""
        # Split out the failing rows by township so that we know which
        # townships are represented in this test's failures
        failing_rows_by_township: typing.Dict[
            typing.Optional[str], typing.List[typing.Dict]
        ] = {}
        for row in self.failing_rows:
            township_code = row.get(TOWNSHIP_FIELD)
            if not failing_rows_by_township.get(township_code):
                failing_rows_by_township[township_code] = []
            failing_rows_by_township[township_code].append(row)

        # These kwargs are shared by all TownshipTestResults, regardless of
        # township code or failure status
        base_kwargs = {
            "name": self.name,
            "table_name": self.table_name,
            "column_name": self.column_name,
            "status": self.status,
            "description": self.description,
            "elapsed_time": self.elapsed_time,
        }

        # If we have any failing rows, split out separate TownshipTestResult
        # objects for each township/row group; otherwise, just create one
        # object with a null township mapped to the passing test
        if failing_rows_by_township:
            return [
                TownshipTestResult(
                    township_code=township_code,
                    failing_rows=rows,
                    **base_kwargs,
                )
                for township_code, rows in failing_rows_by_township.items()
            ]
        return [
            TownshipTestResult(
                township_code=None, failing_rows=[], **base_kwargs
            )
        ]


class TownshipTestResult(TestResult):
    """A variant of TestResult for a test whose results all share the same
    township. Note that township_code is only present in the case of failing
    tests; the township_code will always be None in the case of a passing test,
    since passing tests have no township (or, thinking of it differently,
    passing tests encompass all of the townships)."""

    def __init__(
        self, township_code: typing.Optional[str], *args, **kwargs
    ) -> None:
        self.township_code = township_code
        super().__init__(*args, **kwargs)


class TestCategory:
    """Class to store TestResult objects for a group of dbt tests that share
    the same category. Provides convenience methods for formatting those
    results for output to a workbook and saving them to a cache."""

    # Names of fields that are used for debugging
    possible_debugging_fieldnames = [TEST_NAME_FIELD, DOCS_URL_FIELD]
    # Names of fields that identify the test
    possible_test_metadata_fieldnames = [
        *[SOURCE_TABLE_FIELD, DESCRIPTION_FIELD],
        *possible_debugging_fieldnames,
    ]
    # Names of fields that are used for diagnostics
    possible_diagnostic_fieldnames = [
        TAXYR_FIELD,
        PARID_FIELD,
        CARD_FIELD,
        LAND_LINE_FIELD,
        CLASS_FIELD,
        TOWNSHIP_FIELD,
        WHO_FIELD,
        WEN_FIELD,
    ]
    # The complete set of fixed fields
    possible_fixed_fieldnames = [
        *possible_test_metadata_fieldnames,
        *possible_diagnostic_fieldnames,
    ]

    def __init__(
        self,
        category: str,
        results: typing.Optional[typing.List[TestResult]] = None,
    ) -> None:
        self.category = category
        self.test_results: typing.List[TestResult] = results or []

    def to_dict(self) -> typing.Dict:
        """Serialize the TestCategory object as a dictionary."""
        return {
            "category": self.category,
            "test_results": [result.to_dict() for result in self.test_results],
        }

    @classmethod
    def from_dict(cls, category_dict: typing.Dict) -> "TestCategory":
        """Deserialize a TestCategory object from a dictionary."""
        return TestCategory(
            category=category_dict["category"],
            results=[
                TestResult.from_dict(result_dict)
                for result_dict in category_dict["test_results"]
            ],
        )

    def __repr__(self) -> str:
        num_failing_rows = sum(
            len(result.failing_rows) for result in self.test_results
        )
        return (
            f"TestCategory(category={self.category!r}, "
            f"status={self.status!r}, "
            f"num_tests={len(self.test_results)}, "
            f"num_failing_rows={num_failing_rows})"
        )

    @property
    def fieldnames(self) -> typing.List[str]:
        """Get a list of fieldnames that encapsulates all of the fieldnames
        for all of the rows of tests tracked by this group."""
        fieldnames = []
        for result in self.test_results:
            for fieldname in result.fieldnames:
                if fieldname not in fieldnames:
                    fieldnames.append(fieldname)

        # Remove any fixed fieldnames from the ordered list that are not
        # present in this group
        fixed_field_order = [
            field
            for field in self.possible_fixed_fieldnames
            if field in fieldnames
        ]

        # Reorder the fieldnames so that diagnostic fields are presented in the
        # correct order
        for field in reversed(fixed_field_order):
            fieldnames.insert(0, fieldnames.pop(fieldnames.index(field)))

        return fieldnames

    @property
    def rows(self) -> typing.List[typing.List]:
        """Format the rows of tests tracked by this group, with
        fieldname data excluded. The combination of this property and the
        `fieldnames` property can be used to write to a csv.Writer or
        to an openpyxl.Workbook sheet for the tests tracked by this group."""
        fieldnames = self.fieldnames
        return [
            [row.get(fieldname) for fieldname in fieldnames]
            for result in self.test_results
            for row in result.failing_rows
        ]

    @property
    def status(self) -> Status:
        """Return an aggregate status for this category based on the statuses
        of its TestResult objects."""
        statuses = set(result.status for result in self.test_results)
        # case/match syntax doesn't work with sets, unfortunately
        if statuses == {Status.PASS}:
            return Status.PASS
        if statuses == {Status.WARN}:
            return Status.WARN
        if statuses == {Status.FAIL}:
            return Status.FAIL
        if statuses == {Status.PASS, Status.WARN}:
            return Status.WARN
        if statuses == {Status.PASS, Status.FAIL}:
            return Status.FAIL
        if statuses == {Status.WARN, Status.FAIL}:
            return Status.FAIL
        raise ValueError(f"Unexpected combination of statuses: {statuses}")

    def add_to_workbook(self, workbook: openpyxl.Workbook) -> None:
        """Add a sheet of failed dbt tests to an openpyxl Workbook using data
        from the TestCategory object. Note that we expect the workbook to be
        initialized with write_only=True."""
        # Only add this category to the workbook if it has any failing tests
        if self.status in (Status.PASS, Status.WARN):
            print(
                f"Skipping add_to_workbook for category {self.category} since "
                f"its status is '{self.status!r}'"
            )
            return

        # openpyxl Workbooks are typically created with one untitled active
        # sheet by default, but write-only sheets are an exception to this
        # rule, so we always have to create a new sheet
        sheet = workbook.create_sheet()
        sheet.title = self.category

        # Freeze the header row. The syntax for the freeze_panes attribute is
        # undocumented, but it freezes all rows above and all columns to the
        # left of the given cell identifier. Note that freeze operations must
        # be performed before any data is added to a sheet in a write-only
        # workbook
        data_header_idx = 3  # We have 3 headers; 2 for grouping and 1 for data
        freeze_pane_letter = openpyxl.utils.get_column_letter(
            len(self.test_metadata_fieldnames) + 1
        )
        freeze_pane_number = data_header_idx + 1
        sheet.freeze_panes = f"{freeze_pane_letter}{freeze_pane_number}"

        # Hide columns that are intended for debugging only, so that they don't
        # get in the way of non-technical workbook consumers
        for col_idx in self.debugging_field_indexes:
            sheet.column_dimensions[col_idx].hidden = True

        # Create groupings for columns with a special group header
        bold_font = openpyxl.styles.Font(bold=True)
        italic_font = openpyxl.styles.Font(italic=True)
        title_row, subtitle_row, header_row, merged_cell_range = [], [], [], []
        column_groups = {
            self.test_metadata_field_indexes: {
                "title": "Test description fields",
                "subtitle": "These fields identify a failing test.",
                "fieldnames": self.test_metadata_fieldnames,
                "style": "20 % - Accent4",
                "header_style": "Accent4",
            },
            self.diagnostic_field_indexes: {
                "title": "Unique identifier fields",
                "subtitle": (
                    "These fields identify the row that is failing a test."
                ),
                "fieldnames": self.diagnostic_fieldnames,
                "style": "20 % - Accent1",
                "header_style": "Accent1",
            },
            self.nonfixed_field_indexes: {
                "title": "Problematic fields",
                "subtitle": (
                    "These fields contain values that are causing the test "
                    "to fail."
                ),
                "fieldnames": self.nonfixed_fieldnames,
                "style": "20 % - Accent2",
                "header_style": "Accent2",
            },
        }
        for col_group_indexes, col_metadata in column_groups.items():
            # Sometimes there are no problematic fields for a given test;
            # if this is the case, skip it
            if not col_group_indexes:
                continue

            # Save merged cell info
            for cell_range in [
                f"{col_group_indexes[0]}1:{col_group_indexes[-1]}1",
                f"{col_group_indexes[0]}2:{col_group_indexes[-1]}2",
            ]:
                merged_cell_range.append(cell_range)

            # Fill out and format grouping header
            title_cell = openpyxl.cell.WriteOnlyCell(
                sheet, value=col_metadata["title"]
            )
            title_cell.style = "Note"
            title_cell.font = bold_font
            title_row.append(title_cell)
            # Flesh out the empty title row cells that will be merged later on
            for _ in range(len(col_group_indexes) - 1):
                title_row.append("")

            subtitle_cell = openpyxl.cell.WriteOnlyCell(
                sheet, value=col_metadata["subtitle"]
            )
            subtitle_cell.style = "Note"
            subtitle_cell.font = italic_font
            subtitle_row.append(subtitle_cell)
            for _ in range(len(col_group_indexes) - 1):
                subtitle_row.append("")

            # Fill out and format the data header
            for fieldname in col_metadata["fieldnames"]:
                header_cell = openpyxl.cell.WriteOnlyCell(
                    sheet, value=fieldname
                )
                header_cell.style = col_metadata["header_style"]
                header_cell.font = openpyxl.styles.Font(
                    bold=True, color=openpyxl.styles.colors.WHITE
                )
                header_row.append(header_cell)

        # Initialize the column widths based on the length of values in
        # the header row
        column_widths = {
            openpyxl.utils.get_column_letter(idx + 1): len(fieldname) + 2
            for idx, fieldname in enumerate(self.fieldnames)
        }
        # Iterate the rows to extract data and optionally update the column
        # widths if the length of the cell value exceeds the length of the
        # header value
        data_rows = []
        for row in self.rows:
            data_row = []
            # Start enumeration at 1 since openpyxl columns are 1-indexed
            for col_idx, cell in enumerate(row, 1):
                # Convert row values to string so that Excel doesn't apply
                # autoformatting
                cell_str = str(cell) if cell is not None else ""
                cell = openpyxl.cell.WriteOnlyCell(sheet, value=cell_str)

                # Retrieve the cell style from the column groupings if one
                # exists
                cell_style = None
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                for col_group_indexes, col_metadata in column_groups.items():
                    if column_letter in col_group_indexes:
                        cell_style = col_metadata["style"]
                if cell_style:
                    cell.style = cell_style
                data_row.append(cell)

                # Check if this cell is longer than the longest cell we've seen
                # so far, and adjust the column dimensions accordingly
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                column_widths[column_letter] = max(
                    column_widths.get(column_letter, 0), len(cell_str)
                )

            data_rows.append(data_row)

        # Update column widths so that they fit the longest column
        for (
            column_letter,
            column_width,
        ) in column_widths.items():
            # Pad with an extra two characters to account for the fact that
            # non-monospace fonts do not have consistent character widths,
            # and set a hard limit of 75 characters so no one field takes over
            # the viewport of the spreadsheet
            width = min(column_width + 2, 75)
            sheet.column_dimensions[column_letter].width = width

        # Add filters to fixed columns (i.e. columns that appear in every sheet
        # in the same position)
        fixed_field_indexes = self.fixed_field_indexes
        sheet_max_row_idx = data_header_idx + len(data_rows)
        min_fixed_idx = f"{fixed_field_indexes[0]}{data_header_idx}"
        max_fixed_idx = f"{fixed_field_indexes[-1]}{sheet_max_row_idx}"
        fixed_field_range = f"{min_fixed_idx}:{max_fixed_idx}"
        sheet.auto_filter.ref = fixed_field_range

        # Add the data to the sheet. This should be one of the last steps in
        # this function, since write-only sheets require all formatting to be
        # set before data is added
        sheet.append(title_row)
        sheet.append(subtitle_row)
        sheet.append(header_row)
        for data_row in data_rows:
            sheet.append(data_row)

        # Merge cells in the grouping headers. This approach is a bit of a hack
        # since merged cells are not fully supported in write-only workbooks,
        # hence why it takes place _after_ rows have been added to the sheet
        # whereas most formatting options for write-only workbooks need to
        # happen _before_ data is added. See here for details:
        # https://stackoverflow.com/a/66159254
        for cell_range in merged_cell_range:
            sheet.merged_cells.ranges.add(cell_range)

    @property
    def debugging_fieldnames(self) -> typing.List[str]:
        """Get a list of fieldnames (e.g. ["foo", "bar"]) for fields that
        are used for debugging."""
        return self._filter_for_existing_fieldnames(
            self.possible_debugging_fieldnames
        )

    @property
    def debugging_field_indexes(self) -> tuple:
        """Get a tuple of field indexes (e.g. ["A", "B"]) for fields that
        are used for debugging."""
        return self._filter_for_existing_field_indexes(
            self.possible_debugging_fieldnames
        )

    @property
    def test_metadata_fieldnames(self) -> typing.List[str]:
        """Get a list of fieldnames (e.g. ["foo", "bar"]) for fields that
        are used for identifying tests."""
        return self._filter_for_existing_fieldnames(
            self.possible_test_metadata_fieldnames
        )

    @property
    def test_metadata_field_indexes(self) -> tuple:
        """Get a tuple of field indexes (e.g. ["A", "B"]) for fields that
        are used for identifying tests."""
        return self._filter_for_existing_field_indexes(
            self.possible_test_metadata_fieldnames
        )

    @property
    def diagnostic_fieldnames(self) -> typing.List[str]:
        """Get a list of fieldnames (e.g. ["foo", "bar"]) for fields that
        are used for diagnostics."""
        return self._filter_for_existing_fieldnames(
            self.possible_diagnostic_fieldnames
        )

    @property
    def diagnostic_field_indexes(self) -> tuple:
        """Get a tuple of field indexes (e.g. ["A", "B"]) for fields that
        are used for diagnostics."""
        return self._filter_for_existing_field_indexes(
            self.possible_diagnostic_fieldnames
        )

    @property
    def fixed_fieldnames(self) -> typing.List[str]:
        """Get a list of fieldnames (e.g. ["foo", "bar"]) for fields that
        are fixed (i.e. whose position is always at the start of the sheet,
        for diagnostic purposes)."""
        return self._filter_for_existing_fieldnames(
            self.possible_fixed_fieldnames
        )

    @property
    def fixed_field_indexes(self) -> tuple:
        """Get a list of field indexes (e.g. ["A", "B"]) for fields that
        are fixed (i.e. whose position is always at the start of the sheet,
        for diagnostic purposes)."""
        return self._filter_for_existing_field_indexes(
            self.possible_fixed_fieldnames
        )

    @property
    def nonfixed_fieldnames(self) -> typing.List[str]:
        """Get a list of field names (e.g. ["foo", "bar"]) for fields that
        are nonfixed (i.e. whose position comes after the fixed fields in the
        sheet and are thus variable)."""
        fieldnames = self.fieldnames
        fixed_fieldnames = self.possible_fixed_fieldnames
        return [field for field in fieldnames if field not in fixed_fieldnames]

    @property
    def nonfixed_field_indexes(self) -> tuple:
        """Get a list of field indexes (e.g. ["A", "B"]) for fields that
        are nonfixed (i.e. whose position comes after the fixed fields in the
        sheet and are thus variable)."""
        nonfixed_fieldnames = self.nonfixed_fieldnames
        return self._filter_for_existing_field_indexes(nonfixed_fieldnames)

    def _filter_for_existing_fieldnames(
        self, possible_fieldnames: typing.List[str]
    ) -> typing.List[str]:
        """Helper function to filter a list of `possible_fieldnames` for
        only those fields that exist in the test group, returning the
        names of the fields (e.g. ["foo", "bar"])."""
        existing_fieldnames = self.fieldnames
        return [
            field
            for field in possible_fieldnames
            if field in existing_fieldnames
        ]

    def _filter_for_existing_field_indexes(
        self, possible_fieldnames: typing.List[str]
    ) -> tuple:
        """Helper function to filter a list of `possible_fieldnames` for
        only those fields that exist in the test group, returning the
        indexes of the fields (e.g. ["A", "B"])."""
        existing_fieldnames = self.fieldnames
        return tuple(
            openpyxl.utils.get_column_letter(
                # openpyxl is 1-indexed while the index() method is 0-indexed
                existing_fieldnames.index(field) + 1
            )
            for field in self._filter_for_existing_fieldnames(
                possible_fieldnames
            )
        )


# Help docstring for the command line interface
CLI_DESCRIPTION = """Runs iasWorld data tests and generates an Excel workbook of dbt test failures that can be shared with other teams
for review and correction, along with metadata parquet files that can be uploaded to S3 for long-term result tracking.

This script expects that Python dependencies have been installed from [project.optional-dependencies].dbt_tests.

Expects one required environment variable to be set:

 1. USER: The username of the user who ran the script. This is automatically set during login on Unix systems, but should be set manually elsewhere.

Expects four optional environment variables:

 1. AWS_ATHENA_S3_STAGING_DIR: Location in S3 where Athena query results should be written (defaults to s3://ccao-athena-results-us-east-1)
 2. GIT_SHA: The SHA of the latest git commit (defaults to the output of `git rev-parse`)
 3. GIT_REF: The name of the ref for the latest git commit (defaults to the output of `git rev-parse --abbrev-ref`)
 4. GIT_AUTHOR: The author of the latest git commit (defaults to the output of `git log`)

Outputs three files to the directory specified by the `--output-dir` flag:

  1. `iasworld_test_failures_<date>.xlsx`: Excel workbook to share with other teams
  2. `metadata/test_run/run_year=YYYY/*.parquet`: Metadata about this run, partitioned by year of run and prepped for upload to S3
  3. `metadata/test_run_result/run_year=YYYY/*.parquet`: Metadata about test results (pass, fail, number of failing rows, etc.) in this run,
     partitioned by year of run and prepped for upload to S3

Each sheet in the output workbook represents a category of test, e.g. "valid_range" or "not_null"; each row in a sheet represents a row in
a database that failed a test, with enough metadata that a reader can figure out what conditions caused the test to fail and investigate the root cause."""  # noqa: E501

# Examples to use in the command line interface docstring
CLI_EXAMPLE = """Example usage with no options provided:

    python3 run_iasworld_data_tests.py

Example usage with all options provided:

    AWS_ATHENA_S3_STAGING_DIR=s3://foo-bar-baz/ python3 run_iasworld_data_tests.py
        --output-dir ./iasworld_test_results/
        --township 77
        --no-use-cached

Example usage to filter for multiple townships:

    python3 run_iasworld_data_tests.py --township 76 77

Example usage to skip running tests, and instead reuse results from a previous run:

    python3 run_iasworld_data_tests.py --use-cached

"""  # noqa: E501


def main() -> None:
    """Entrypoint to this script. Runs dbt tests and writes artifacts
    to the output directory with metadata about test results."""

    parser = argparse.ArgumentParser(
        description=CLI_DESCRIPTION,
        epilog=CLI_EXAMPLE,
        # Parse the description and epilog as raw text so that newlines
        # get preserved
        formatter_class=argparse.RawTextHelpFormatter,
    )

    parser.add_argument(
        "--output-dir",
        required=False,
        help=(
            "The directory to which output artifacts should be written; "
            "if the directory does not exist, it will be created. Defaults to "
            "'./iasworld_test_results_<date>/'."
        ),
    )
    parser.add_argument(
        "--township",
        required=False,
        nargs="*",
        help=(
            "One or more optional township codes which will be used to filter "
            "results. Can be provided with multiple space-separated values "
            "to select multiple townships. Defaults to all townships, "
            "including null townships (which typically indicate invalid PINs)."
        ),
    )
    parser.add_argument(
        "--use-cached",
        action=argparse.BooleanOptionalAction,
        required=False,
        help=(
            "Toggle using cached results from the most recent run. Useful when debugging "
            "transformation steps. Defaults to False."
        ),
    )
    parser.add_argument(
        "--skip-artifacts",
        action=argparse.BooleanOptionalAction,
        required=False,
        help=(
            "Just run tests and skip the step that parses test output. "
            "Ignored if --use-cached is set, since --use-cached implies "
            "that the script should skip running tests"
        ),
    )
    parser.add_argument(
        "--defer",
        action=argparse.BooleanOptionalAction,
        required=False,
        default=False,
        help=(
            "Same as the dbt --defer option, resolves unselected nodes by "
            "deferring to the manifest within the --state directory"
        ),
    )
    parser.add_argument(
        "--state",
        required=False,
        help=(
            "Same as the dbt --state option, use this state directory for "
            "deferral"
        ),
    )
    parser.add_argument(
        *constants.SELECT_ARGUMENT_ARGS, **constants.SELECT_ARGUMENT_KWARGS
    )
    parser.add_argument(
        *constants.SELECTOR_ARGUMENT_ARGS, **constants.SELECTOR_ARGUMENT_KWARGS
    )
    parser.add_argument(
        *constants.TARGET_ARGUMENT_ARGS, **constants.TARGET_ARGUMENT_KWARGS
    )

    args = parser.parse_args()

    output_dir = args.output_dir
    townships = args.township if args.township else tuple()
    use_cached = args.use_cached
    skip_artifacts = args.skip_artifacts
    defer = args.defer
    state = args.state
    select = args.select
    selector = args.selector
    target = args.target

    run_results_filepath = os.path.join("target", "run_results.json")
    manifest_filepath = os.path.join("target", "manifest.json")

    date_today = datetime.datetime.today().strftime("%Y-%m-%d")
    if output_dir is None:
        output_dir = f"iasworld_test_results_{date_today}"

    if (not defer and state) or (defer and not state):
        raise ValueError("--defer and --state must be used together")

    select_args = ["--selector", "select_data_test_iasworld"]
    if select:
        select_args = ["--select", *select]
    if selector:
        select_args = ["--selector", selector]

    if use_cached:
        test_cache_path = get_test_cache_path(
            run_results_filepath, manifest_filepath, townships
        )
        if os.path.isfile(test_cache_path):
            print(f"Loading test results from cache at {test_cache_path}")
            test_categories = get_test_categories_from_file(test_cache_path)
        else:
            raise ValueError(
                f"Test cache not found at {test_cache_path}, try rerunning "
                "without --use-cached"
            )
    else:
        print("Running tests")
        dbt_run_args = ["test", "--target", target, *select_args]
        if not skip_artifacts:
            dbt_run_args.append("--store-failures")
        if defer and state:
            dbt_run_args += ["--defer", "--state", state]

        print(f"> dbt {' '.join(dbt_run_args)}")
        dbt_test_result = DBT.invoke(dbt_run_args)

        if dbt_test_result.exception is not None:
            raise dbt_test_result.exception

        if any(
            result.status == TestStatus.Error
            for result in getattr(dbt_test_result.result, "results", [])
        ):
            # No need to report the exception, since the dbt process
            # will have printed it already
            raise ValueError("Quitting due to error in dbt test run")

        if skip_artifacts:
            print("Skipping artifact generation since --skip-artifacts is set")
            return

        print("Loading test results from Athena")
        test_categories = get_test_categories_from_athena(
            run_results_filepath, manifest_filepath, townships
        )

        new_test_cache_path = get_test_cache_path(
            run_results_filepath, manifest_filepath, townships
        )
        print(f"Saving test results to the cache at {new_test_cache_path}")
        save_test_categories_to_file(test_categories, new_test_cache_path)

    print("Generating the output workbook")
    # It's important to use a write-only workbook here because otherwise
    # the metadata required to store cell info about a large number of failing
    # tests can cause the process to run out of memory
    workbook = openpyxl.Workbook(write_only=True)
    for test_category in test_categories:
        print(f"Adding sheet for {test_category.category}")
        test_category.add_to_workbook(workbook)

    pathlib.Path(output_dir).mkdir(exist_ok=True)
    workbook_filepath = os.path.join(
        output_dir, f"iasworld_test_failures_{date_today}.xlsx"
    )
    workbook.save(workbook_filepath)
    print(f"Output workbook saved to {workbook_filepath}")

    # Get run metadata from the environment
    try:
        run_by = os.environ["USER"]
    except KeyError:
        raise ValueError("USER env variable must be set")

    git_sha = (
        os.environ["GIT_SHA"]
        if os.getenv("GIT_SHA")
        else subprocess.getoutput("git rev-parse HEAD")
    )
    git_ref = (
        os.environ["GIT_REF"]
        if os.getenv("GIT_REF")
        else subprocess.getoutput("git rev-parse --abbrev-ref HEAD")
    )
    git_author = (
        os.environ["GIT_AUTHOR"]
        if os.getenv("GIT_AUTHOR")
        else subprocess.getoutput("git log -1 --pretty=format:'%an <%ae>'")
    )

    # Generate and save metadata tables as parquet
    test_run_metadata = TestRunMetadata.create(
        run_results_filepath, run_by, git_sha, git_ref, git_author
    )
    test_run_result_metadata_list = TestRunResultMetadata.create_list(
        test_categories, run_results_filepath
    )
    test_run_failing_row_metadata_list = TestRunFailingRowMetadata.create_list(
        test_categories, run_results_filepath
    )
    run_date = get_run_date_from_run_results(run_results_filepath)
    run_id = get_run_id_from_run_results(run_results_filepath)

    for metadata_list, tablename, partition_cols in [
        ([test_run_metadata], "test_run", ["run_year"]),
        (test_run_result_metadata_list, "test_run_result", ["run_year"]),
        (
            test_run_failing_row_metadata_list,
            "test_run_failing_row",
            ["run_year"],
        ),
    ]:
        if not metadata_list:
            print(f"{tablename} is empty, skipping metadata output")
            continue

        table = pa.Table.from_pylist(
            [meta_obj.to_dict() for meta_obj in metadata_list],  # type: ignore
        )
        metadata_root_path = os.path.join(output_dir, "metadata", tablename)
        pyarrow.parquet.write_to_dataset(
            table,
            metadata_root_path,
            partition_cols,
            basename_template="%s_%s_{i}.parquet" % (run_date, run_id),
        )
        print(f"{tablename} metadata saved to {metadata_root_path}/")


@dataclasses.dataclass
class TestRunMetadata:
    """Metadata object storing information about a test run."""

    run_id: str
    run_date: str
    run_year: str  # Separate from run_date for partitioning
    run_by: str
    elapsed_time: decimal.Decimal
    var_year_start: str
    var_year_end: str
    git_sha: str
    git_ref: str
    git_author: str

    @classmethod
    def create(
        cls,
        run_results_filepath: str,
        run_by: str,
        git_sha: str,
        git_ref: str,
        git_author: str,
    ) -> "TestRunMetadata":
        """Generate a TestRunMetadata object from a filepath to a
        run_results.json file."""
        run_id = get_run_id_from_run_results(run_results_filepath)
        run_date = get_run_date_from_run_results(run_results_filepath)
        run_year = run_date[:4]
        elapsed_time = get_key_from_run_results(
            "elapsed_time", run_results_filepath
        )

        # Extract dbt vars
        run_vars = get_key_from_run_results("args", run_results_filepath)[
            "vars"
        ]
        var_year_start = run_vars.get("data_test_iasworld_year_start")
        var_year_end = run_vars.get("data_test_iasworld_year_end")

        # If dbt vars weren't set on the command line, the defaults won't exist
        # in run_results.json, so we have to parse them from the dbt project
        # config
        if not var_year_start or not var_year_end:
            with open("dbt_project.yml") as project_fobj:
                project = yaml.safe_load(project_fobj)
            var_year_start = (
                var_year_start
                or project["vars"]["data_test_iasworld_year_start"]
            )
            var_year_end = (
                var_year_end or project["vars"]["data_test_iasworld_year_end"]
            )

        return cls(
            run_id=run_id,
            run_year=run_year,
            run_date=run_date,
            run_by=run_by,
            elapsed_time=elapsed_time,
            var_year_start=var_year_start,
            var_year_end=var_year_end,
            git_sha=git_sha,
            git_ref=git_ref,
            git_author=git_author,
        )

    def to_dict(self) -> typing.Dict:
        return dataclasses.asdict(self)


@dataclasses.dataclass
class TestRunResultMetadata:
    """Metadata object storing aggregated information about township-level
    test results in a run."""

    run_id: str
    run_year: str  # Duplicated with TestRunMetadata for partitioning
    test_name: str
    table_name: str
    column_name: typing.Optional[str]
    category: str
    description: str
    township_code: typing.Optional[str]
    status: str  # Serialize Status enum to str for output to parquet
    elapsed_time: decimal.Decimal
    num_failing_rows: int

    @classmethod
    def create_list(
        cls,
        test_categories: typing.List[TestCategory],
        run_results_filepath: str,
    ) -> typing.List["TestRunResultMetadata"]:
        """Generate a list of TestRunMetadata object from a list of
        TestCategory objects representing the categories in the run and a
        filepath to a run_results.json file."""
        run_id = get_run_id_from_run_results(run_results_filepath)
        run_date = get_run_date_from_run_results(run_results_filepath)
        run_year = run_date[:4]

        return [
            TestRunResultMetadata(
                run_id=run_id,
                run_year=run_year,
                test_name=township_result.name,
                table_name=township_result.table_name,
                column_name=township_result.column_name,
                category=test_category.category,
                description=township_result.description,
                township_code=township_result.township_code,
                status=township_result.status.value,
                elapsed_time=township_result.elapsed_time,
                num_failing_rows=len(township_result.failing_rows),
            )
            for test_category in test_categories
            for test_result in test_category.test_results
            for township_result in test_result.split_by_township()
        ]

    def to_dict(self) -> typing.Dict:
        return dataclasses.asdict(self)


@dataclasses.dataclass
class TestRunFailingRowMetadata:
    """Metadata object storing information about row-level individual test
    failures in a run."""

    # Fields that identify the run
    run_id: str
    run_year: str  # Duplicated with TestRunMetadata for partitioning
    # Fields that identify the test
    test_name: str
    table_name: str
    column_name: typing.Optional[str]
    category: str
    description: str
    # Fields that identify the failing row. Some of these can occasionally
    # be arrays for tests that query multiple rows (e.g. uniqueness tests)
    # so for consistency we set them to always be arrays, even when there
    # is only one value
    parid: typing.Optional[str]
    taxyr: typing.Optional[str]
    card: typing.Optional[typing.List[int]]
    lline: typing.Optional[typing.List[int]]
    class_: typing.Optional[typing.List[str]]
    township_code: typing.Optional[str]
    who: typing.Optional[typing.List[str]]
    wen: typing.Optional[typing.List[str]]
    # Since the problematic fields can vary so widely, we store them as a
    # JSON blob
    problematic_fields: typing.Dict

    @classmethod
    def create_list(
        cls,
        test_categories: typing.List[TestCategory],
        run_results_filepath: str,
    ) -> typing.List["TestRunFailingRowMetadata"]:
        """Generate a list of TestRunFailingRowMetadata object from a list of
        TestCategory objects representing the categories in the run and a
        filepath to a run_results.json file."""
        run_id = get_run_id_from_run_results(run_results_filepath)
        run_date = get_run_date_from_run_results(run_results_filepath)
        run_year = run_date[:4]

        def value_to_list(value):
            """Tiny helper function to convert not-null column values to lists.
            Useful in cases where a column can be either a scalar, a list,
            or a null value, in which cases we want the output to always be
            either a null value or a list."""
            if value is None or type(value) is list:
                return value
            return [value]

        return [
            TestRunFailingRowMetadata(
                run_id=run_id,
                run_year=run_year,
                test_name=township_result.name,
                table_name=township_result.table_name,
                column_name=township_result.column_name,
                category=test_category.category,
                description=township_result.description,
                township_code=township_result.township_code,
                parid=failing_row.get(PARID_FIELD),
                taxyr=failing_row.get(TAXYR_FIELD),
                card=value_to_list(failing_row.get(CARD_FIELD)),
                lline=value_to_list(failing_row.get(LAND_LINE_FIELD)),
                class_=value_to_list(failing_row.get(CLASS_FIELD)),
                who=value_to_list(failing_row.get(WHO_FIELD)),
                wen=value_to_list(failing_row.get(WEN_FIELD)),
                problematic_fields={
                    key: val
                    for key, val in failing_row.items()
                    # Use possible_fixed_fieldnames to avoid having to
                    # recompute the exact fixed fieldnames on every iteration
                    # (we could also solve this by expanding out the list
                    # comprehension, but for now this is easier)
                    if key not in test_category.possible_fixed_fieldnames
                },
            )
            for test_category in test_categories
            for test_result in test_category.test_results
            for township_result in test_result.split_by_township()
            for failing_row in township_result.failing_rows
            if township_result.status == Status.FAIL
        ]

    def to_dict(self) -> typing.Dict:
        output_data = dataclasses.asdict(self)
        # Serialize the "class" attribute to a more human-friendly name
        output_data["class"] = output_data.pop("class_")
        # Dump the problematic fields to string, since parquet can't handle
        # the notion of an untyped JSON object
        output_data["problematic_fields"] = json.dumps(
            output_data["problematic_fields"]
        )
        return output_data


def get_key_from_run_results(
    key: str, run_results_filepath: str
) -> typing.Any:
    """Given a path to a run_results.json file, return a key that's represented
    in the run metadata."""
    with open(run_results_filepath) as run_results_fobj:
        run_results = json.load(run_results_fobj)

    return run_results[key]


def get_run_id_from_run_results(run_results_filepath: str) -> str:
    """Given a path to a run_results.json file, return a string representation
    of the invocation ID of the run."""
    metadata = get_key_from_run_results("metadata", run_results_filepath)
    return metadata["invocation_id"]


def get_run_date_from_run_results(run_results_filepath: str) -> str:
    """Given a path to a run_results.json file, return a string representation
    of the date of the run formatted as YYYY-MM-DD."""
    metadata = get_key_from_run_results("metadata", run_results_filepath)
    run_dt_str = metadata["generated_at"]
    run_dt = datetime.datetime.strptime(run_dt_str, "%Y-%m-%dT%H:%M:%S.%fZ")
    return run_dt.strftime("%Y-%m-%d")


def get_test_cache_path(
    run_results_filepath: str,
    manifest_filepath: str,
    townships: typing.Tuple[str],
) -> str:
    """Return the path to the cache where test results are stored, or an
    empty string if no cache exists yet. The `run_results_filepath` and
    `manifest_filepath` are used to generated a hash key that uniquely defines
    the cache key for a given test run."""
    if not os.path.isfile(run_results_filepath) or not os.path.isfile(
        manifest_filepath
    ):
        return ""

    with open(run_results_filepath, "rb") as run_results_file:
        run_results_hash = hashlib.md5(run_results_file.read()).hexdigest()

    with open(manifest_filepath, "rb") as manifest_file:
        manifest_hash = hashlib.md5(manifest_file.read()).hexdigest()

    base_filename = f"run_{run_results_hash}_manifest_{manifest_hash}"
    if townships:
        base_filename += f"_township_{'_'.join(townships)}"

    return os.path.join(TEST_CACHE_DIR, f"{base_filename}.json")


def get_test_categories_from_file(
    file_path: str,
) -> typing.List[TestCategory]:
    """Load a list of TestCategory objects from a cache located at
    `file_path`."""
    with open(file_path) as cache_file:
        test_category_dicts = json.load(cache_file, use_decimal=True)
    return [
        TestCategory.from_dict(category_dict)
        for category_dict in test_category_dicts
    ]


def save_test_categories_to_file(
    test_categories: typing.List[TestCategory], file_path: str
) -> None:
    """Save a list of TestCategory objects to a cache located at
    `file_path`."""
    test_category_dicts = [
        test_category.to_dict() for test_category in test_categories
    ]
    os.makedirs(TEST_CACHE_DIR, exist_ok=True)
    with open(file_path, "w") as cache_file:
        json.dump(test_category_dicts, cache_file, use_decimal=True)


def get_test_categories_from_athena(
    run_results_filepath: str,
    manifest_filepath: str,
    townships: typing.Tuple[str],
) -> typing.List[TestCategory]:
    """Load a list of TestCategory objects by querying Athena for
    test results generated from a `dbt test --store-failures` call,
    optionally filtering results by township."""
    with open(run_results_filepath) as run_results_fobj:
        run_results = json.load(run_results_fobj)

    with open(manifest_filepath) as manifest_fobj:
        manifest = json.load(manifest_fobj)

    test_categories = get_test_categories(run_results, manifest, townships)
    if not test_categories:
        raise ValueError(f"{run_results_filepath} contains no test results")

    return test_categories


def get_test_categories(
    run_results: typing.Dict,
    manifest: typing.Dict,
    townships: typing.Tuple[str],
) -> typing.List[TestCategory]:
    """Given two artifacts from a `dbt test --store-failures` call (a
    run_results.json file dict and a manifest.json file dict) and an optional
    township filter, generates a list of TestCategory objects storing the
    results of the tests."""
    # Define a cursor with unload=True to output query results as parquet.
    # This is particularly important when selecting aggregated columns,
    # which are deserialized incorrectly by regular cursors
    cursor = pyathena.connect(
        s3_staging_dir=AWS_ATHENA_S3_STAGING_DIR,
        region_name="us-east-1",
        cursor_class=pyathena.arrow.cursor.ArrowCursor,
    ).cursor(unload=True)

    tests_by_category: typing.Dict[str, TestCategory] = {}

    for run_result in run_results["results"]:
        unique_id = run_result["unique_id"]
        node = manifest["nodes"].get(unique_id)
        if node is None:
            raise ValueError(f"Missing dbt manifest node with id {unique_id}")

        test_name = node["name"]
        status = run_result["status"]
        execution_time = run_result["execution_time"]

        meta = node.get("meta", {})
        category = get_category_from_node(node)
        tablename = get_tablename_from_node(node)
        column_name = get_column_name_from_node(node)
        test_description = meta.get("description")

        # Basic attrs for the test result that apply whether or not the test
        # failed
        base_result_kwargs = {
            "name": test_name,
            "table_name": tablename,
            "column_name": column_name,
            "description": test_description,
            "elapsed_time": execution_time,
        }

        if not tests_by_category.get(category):
            tests_by_category[category] = TestCategory(category=category)

        if status == Status.PASS.value:
            test_result = TestResult(
                status=Status.PASS, failing_rows=[], **base_result_kwargs
            )

        elif status in (Status.FAIL.value, Status.WARN.value):
            # Link to the test's page in the dbt docs, for debugging
            test_docs_url = f"{DOCS_URL_PREFIX}/{unique_id}"

            # Get the fully-qualified name of the table that stores failures
            # for this test so that we can query it
            test_results_relation_name = node.get("relation_name")
            if test_results_relation_name is None:
                raise ValueError(
                    f"Missing relation_name attribute for test {test_name}. "
                    "Did you run `dbt test` with the --store-failures flag?"
                )

            print(f"Querying failed rows from {test_results_relation_name}")
            # Athena SHOW COLUMNS doesn't allow double quoted tablenames
            relation_name_unquoted = test_results_relation_name.replace(
                '"', "`"
            )
            cursor.execute(f"show columns in {relation_name_unquoted}")
            # SHOW COLUMNS often returns field names with trailing whitespace
            fieldnames = [row[0].strip() for row in cursor]

            # Construct the query to select the test results that were stored
            # in Athena. We want to rehydrate a few fields like township_code
            # and class without requiring that all tests select them, so
            # check if those fields are missing and join against the correct
            # source table in an attempt to rehydrate the missing columns.
            #
            # The conditionals that follow are a bit ugly, but they reflect the
            # fact that we need to rehydrate different columns from different
            # source tables, and that those source tables can differ based on
            # whether the base model is keyed by card or parcel. In the future,
            # we might consider simplifying this by joining to a helper view
            # with a universe of township and class codes split out by table,
            # e.g.:
            #
            #    pin  | card | taxyr | township_code |  table   | class
            #   ----- | ---- | ----- | ------------- | -------- | -----
            #   12345 | 1    | 2024  | 10            | land     | 200
            #   12345 | 1    | 2024  | 10            | dweldat  | 211
            test_results_select = "select test_results.*"
            test_results_join = ""
            # Format the townships tuple as a SQL array for filtering
            townships_sql = (
                "(" + ",".join(f"'{code}'" for code in townships) + ")"
            )
            # The correct reference for the township column varies
            # depending on whether the original test selected a
            # township field or not. If the township column has been
            # selected, we can always reference it in the filter; otherwise,
            # we need to wait to see if we can rehydrate the township based
            # on the parid and taxyr
            test_results_filter = (
                f" where test_results.{TOWNSHIP_FIELD} in {townships_sql}"
                if townships and TOWNSHIP_FIELD in fieldnames
                else ""
            )
            # We need parid and taxyr at minimum in order to rehydrate any
            # missing fields
            if PARID_FIELD in fieldnames and TAXYR_FIELD in fieldnames:
                if TOWNSHIP_FIELD not in fieldnames:
                    test_results_select += f", leg.user1 AS {TOWNSHIP_FIELD}"
                    test_results_join += f"""
                        left join iasworld.legdat as leg
                            on leg.{PARID_FIELD} = test_results.{PARID_FIELD}
                            and leg.{TAXYR_FIELD} = test_results.{TAXYR_FIELD}
                            and leg.cur = 'Y'
                            and leg.deactivat is null
                    """
                    if townships:
                        test_results_filter = (
                            f" where leg.user1 in {townships_sql}"
                        )
                if CLASS_FIELD not in fieldnames:
                    if (
                        LAND_LINE_FIELD in fieldnames
                        or CARD_FIELD in fieldnames
                    ):
                        # Figure out the right table and key to join on in
                        # order to query the class
                        card_join_table = "dweldat"
                        if tablename in ["comdat", "land", "oby"]:
                            card_join_table = tablename

                        card_field = (
                            LAND_LINE_FIELD
                            if tablename == "land"
                            and LAND_LINE_FIELD in fieldnames
                            else CARD_FIELD
                        )

                        test_results_select += f", card.class AS {CLASS_FIELD}"
                        test_results_join += f"""
                            left join iasworld.{card_join_table} as card
                                on card.{PARID_FIELD} =
                                    test_results.{PARID_FIELD}
                                and card.{TAXYR_FIELD} =
                                    test_results.{TAXYR_FIELD}
                                and card.{card_field} =
                                    test_results.{card_field}
                                and card.cur = 'Y'
                                and card.deactivat is null
                        """
                    else:
                        test_results_select += f", par.class AS {CLASS_FIELD}"
                        test_results_join += f"""
                            left join iasworld.pardat as par
                                on par.{PARID_FIELD} =
                                    test_results.{PARID_FIELD}
                                and par.{TAXYR_FIELD} =
                                    test_results.{TAXYR_FIELD}
                                and par.cur = 'Y'
                                and par.deactivat is null
                        """

            test_results_query = (
                f"{test_results_select} "
                f"from {test_results_relation_name} as test_results "
                f"{test_results_join}"
                f"{test_results_filter}"
            )
            # Use the cursor with unload=True in this query, since otherwise
            # aggregated columns can be deserialized incorrectly
            cursor.execute(test_results_query)
            query_results = cursor.as_arrow().to_pylist()
            if len(query_results) == 0:
                msg = (
                    f"Test {test_name} has status {status!r} but no failing "
                    "rows in Athena"
                )
                if townships:
                    # Missing rows are most likely due to the township filter,
                    # so print a warning and skip
                    print(msg + ", possibly due to the township filter")
                    test_result = TestResult(
                        status=Status.PASS,
                        failing_rows=[],
                        **base_result_kwargs,
                    )
                else:
                    # If there's no township filter, the lack of rows indicates
                    # an unexpected error
                    raise ValueError(msg)
            else:
                # Add custom fields to query results that we don't expect to be
                # included in the response
                failing_rows = [
                    {
                        TEST_NAME_FIELD: test_name,
                        DESCRIPTION_FIELD: test_description,
                        DOCS_URL_FIELD: test_docs_url,
                        SOURCE_TABLE_FIELD: tablename,
                        **row,
                    }
                    for row in query_results
                ]
                test_result = TestResult(
                    status=Status(status),
                    failing_rows=failing_rows,
                    **base_result_kwargs,
                )

        else:
            raise ValueError(
                f"Got unrecognized status '{status}' for node {unique_id} "
                "in dbt run results"
            )

        tests_by_category[category].test_results.append(test_result)

    # Now that we've accumulated all of the test results and they are grouped
    # into categories, we no longer need the category key in the dict, so
    # transform the output into a list
    return list(tests_by_category.values())


def get_category_from_node(node: typing.Dict) -> str:
    """Given a Node for a test extracted from a dbt manifest, return the
    category that the test should go in."""
    if meta_category := node.get("meta", {}).get("category"):
        return meta_category

    for dependency_macro in node["depends_on"]["macros"]:
        # Macro names in the DAG are fully qualified following the pattern
        # `macro.<project_name>.<macro_name>`, so remove the prefix to extract
        # just the name of the macro
        cleaned_macro_name = dependency_macro.split(".")[-1]
        if custom_test_name := TEST_CATEGORIES.get(cleaned_macro_name):
            return custom_test_name
        # Custom generic tests are always formatted like test_<generic_name>
        if cleaned_macro_name.startswith("test_"):
            return cleaned_macro_name.split("test_")[-1]

    return DEFAULT_TEST_CATEGORY


def get_tablename_from_node(node: typing.Dict) -> str:
    """Given a Node for a test extracted from a dbt manifest, return the name
    of the table that the test is testing."""
    if meta_tablename := node.get("meta", {}).get("table_name"):
        # If meta.table_name is set, treat it as an override
        return meta_tablename

    # Search for the model that is implicated in this test via the
    # test_metadata.kwargs.model attribute. Note that it is common to use the
    # elements in the depends_on array for this purpose, but this approach
    # is fraught, since the order of parents for tests with multiple
    # dependencies is not clear and can differ:
    # https://github.com/dbt-labs/dbt-core/issues/6746#issuecomment-1829860236
    test_metadata = node.get("test_metadata", {})
    model_getter_str = test_metadata.get("kwargs", {}).get("model")
    if not model_getter_str:
        raise ValueError(
            "Can't infer tablename: Missing `test_metadata.kwargs.model`"
            f"attribute for test {node['name']}. You may need to add a "
            "`meta.table_name` attribute to the test config to manually "
            "specify the tablename"
        )

    # The test_metadata.kwargs.model attribute is formatted as a Jinja template
    # call to the get_where_subquery macro, so we need to extract the ref or
    # source tablename from that call
    ref_match = re.search(r"ref\('(.+)'\)", model_getter_str)
    if ref_match is not None:
        fq_model_name = ref_match.group(1)
        return fq_model_name.split(".")[-1]

    source_match = re.search(r"source\('iasworld', '(.+)'\)", model_getter_str)
    if source_match is not None:
        return source_match.group(1)

    raise ValueError(
        "Can't infer tablename: Failed to parse model name from "
        f'`test_metadata.kwargs.model` attribute "{model_getter_str}" '
        f" for test \"{node['name']}\". Inspect the dbt manifest file "
        "for more information"
    )


def get_column_name_from_node(node: typing.Dict) -> typing.Optional[str]:
    """Given a Node for a test extracted from a dbt manifest, return the name
    of the column that the test is testing. Note that the column name is not
    always set, e.g. for tests that are defined on a table instead of on
    a column, so the return value can be None."""
    if meta_column_name := node.get("meta", {}).get("column_name"):
        # If meta.column_name is set, treat it as an override
        return meta_column_name

    return node.get("column_name")


if __name__ == "__main__":
    main()
