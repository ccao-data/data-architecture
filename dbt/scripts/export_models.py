# Export dbt models to Excel files.
#
# Run `python scripts/export_models.py --help` for details.

import argparse
import contextlib
import io
import json
import os
import pathlib
import shutil

import pandas as pd
import pyathena
from dbt.cli.main import dbtRunner
from openpyxl.styles import Alignment
from openpyxl.styles.numbers import FORMAT_NUMBER
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

DBT = dbtRunner()
CLI_DESCRIPTION = """Export dbt models to Excel files.

Expects dependencies from requirements.txt (dbt dependencies) and scripts/requirements.export_models.txt (script dependencies) be installed.

A few configuration values can be set on any model to support exporting:

    * config.meta.export_name (optional): The base name of the output file that will be generated. File extensions are not necessary since all
      models are exported as .xlsx files. If unset, defaults to the name of the model.

    * config.meta.export_template (optional): The filename of an Excel template to use when exporting a model. Templates should be stored in the
      export/templates/ directory and should include header rows. If unset, will search for a template with the same name as the model; if no
      template is found, defaults to a simple layout with filterable columns and striped rows.
"""  # noqa: E501
CLI_EXAMPLE = """Example usage to output the qc_report_town_close report for Hyde Park township:

    python scripts/export_models.py --select tag:qc_report_town_close --where "township_code = '70'"
"""  # noqa: E501


def main():
    parser = argparse.ArgumentParser(
        description=CLI_DESCRIPTION,
        epilog=CLI_EXAMPLE,
        # Parse the description and epilog as raw text so that newlines
        # get preserved
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument(
        "--select",
        required=True,
        nargs="*",
        help="One or more dbt select statements to use for filtering models",
    )
    parser.add_argument(
        "--target",
        required=False,
        default="dev",
        help="dbt target to use for querying model data, defaults to 'dev'",
    )
    parser.add_argument(
        "--rebuild",
        action=argparse.BooleanOptionalAction,
        default=False,
        help="Rebuild models before exporting",
    )
    parser.add_argument(
        "--where",
        required=False,
        help="SQL expression representing a WHERE clause to filter models",
    )

    args = parser.parse_args()
    target = args.target
    select = args.select
    rebuild = args.rebuild
    where = args.where

    if rebuild:
        dbt_run_args = [
            "run",
            "--target",
            target,
            "--select",
            *select,
        ]
        print("Rebuilding models")
        print(f"> dbt {' '.join(dbt_run_args)}")
        dbt_run_result = DBT.invoke(dbt_run_args)
        if not dbt_run_result.success:
            print("Encountered error in `dbt run` call")
            raise ValueError(dbt_run_result.exception)

    print("Listing models to select for export")
    dbt_list_args = [
        "--quiet",
        "list",
        "--target",
        target,
        "--resource-types",
        "model",
        "--output",
        "json",
        "--output-keys",
        "name",
        "config",
        "relation_name",
        "--select",
        *select,
    ]
    print(f"> dbt {' '.join(dbt_list_args)}")
    dbt_output = io.StringIO()
    with contextlib.redirect_stdout(dbt_output):
        dbt_list_result = DBT.invoke(dbt_list_args)

    if not dbt_list_result.success:
        print("Encountered error in `dbt list` call")
        raise ValueError(dbt_list_result.exception)

    # Output is formatted as a list of newline-separated JSON objects
    models = [
        json.loads(model_dict_str)
        for model_dict_str in dbt_output.getvalue().split("\n")
        # Filter out empty strings caused by trailing newlines
        if model_dict_str
    ]

    if not models:
        raise ValueError(f"No models found for the --select value '{select}'")

    print(
        "The following models will be exported: "
        f"{', '.join(model['name'] for model in models)}"
    )

    conn = pyathena.connect(
        s3_staging_dir=os.getenv(
            "AWS_ATHENA_S3_STAGING_DIR",
            "s3://ccao-dbt-athena-results-us-east-1",
        ),
        region_name=os.getenv("AWS_ATHENA_REGION_NAME", "us-east-1"),
    )

    for model in models:
        # Extract useful model metadata from the columns we queried in
        # the `dbt list` call above
        model_name = model["name"]
        relation_name = model["relation_name"]
        export_name = model["config"]["meta"].get("export_name") or model_name
        template = (
            model["config"]["meta"].get("export_template")
            or f"{model_name}.xlsx"
        )

        # Define inputs and outputs for export based on model metadata
        template_path = os.path.join("export", "templates", template)
        template_exists = os.path.isfile(template_path)
        output_path = os.path.join("export", "output", f"{export_name}.xlsx")

        print(f"Querying data for model {model_name}")
        query = f"SELECT * FROM {relation_name}"
        if where:
            query += f" WHERE {where}"
        print(f"> {query}")
        model_df = pd.read_sql(query, conn)

        # Delete the output file if one already exists
        pathlib.Path(output_path).unlink(missing_ok=True)

        if template_exists:
            print(f"Using template file at {template_path}")
            shutil.copyfile(template_path, output_path)
        else:
            print("No template file exists; creating a workbook from scratch")

        writer_kwargs = (
            {"mode": "a", "if_sheet_exists": "overlay"}
            if template_exists
            else {}
        )
        with pd.ExcelWriter(
            output_path, engine="openpyxl", **writer_kwargs
        ) as writer:
            sheet_name = "Sheet1"
            model_df.to_excel(
                writer,
                sheet_name=sheet_name,
                header=False if template_exists else True,
                index=False,
                startrow=1 if template_exists else 0,
            )
            sheet = writer.sheets[sheet_name]

            # Add a table for data filtering. Only do this if the result set
            # is not empty, because otherwise the empty table will make
            # the Excel workbook invalid
            if model_df.empty:
                print(
                    "Skipping formatting for output workbook since result set "
                    "is empty"
                )
            else:
                table = Table(
                    displayName="Query_Results",
                    ref=(
                        f"A1:{get_column_letter(sheet.max_column)}"
                        f"{str(sheet.max_row)}"
                    ),
                )
                table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium11", showRowStripes=True
                )
                sheet.add_table(table)

                # If a parid column exists, format it explicitly as a number to
                # avoid Excel converting it to scientific notation when a user
                # edits the cell
                if "parid" in model_df or "pin" in model_df:
                    parid_field = "parid" if "parid" in model_df else "pin"
                    parid_index = model_df.columns.get_loc(parid_field)
                    # Skip header row when applying formatting. We need to
                    # catch the special case where there is only one row, or
                    # else we will iterate the _cells_ in that row instead of
                    # the row when slicing it from 2 : max_row
                    non_header_rows = (
                        [sheet[2]]
                        if sheet.max_row == 2
                        else sheet[2 : sheet.max_row]
                    )
                    for row in non_header_rows:
                        row[parid_index].number_format = FORMAT_NUMBER
                        # Left align since PINs do not actually need to be
                        # compared by order of magnitude the way that numbers
                        # do
                        row[parid_index].alignment = Alignment(
                            horizontal="left"
                        )

                # Apply any column formatting that was configured
                format_config = model["config"]["meta"].get(
                    "export_format", {}
                )
                if column_configs := format_config.get("columns"):
                    for column_config in column_configs:
                        # Set horizontal alignment if config is present
                        if horiz_align_dir := column_config.get(
                            "horizontal_align"
                        ):
                            horizontal_alignment = Alignment(
                                horizontal=horiz_align_dir
                            )
                            col_letter = column_config.get("index")
                            if col_letter is None:
                                raise ValueError(
                                    "'index' attribute is required when "
                                    "'horizontal_align' is set on "
                                    "export_format.columns config for "
                                    f"model {model_name}"
                                )
                            idx = column_index_from_string(col_letter) - 1
                            # Skip header row
                            for row in sheet[2 : sheet.max_row]:
                                row[idx].alignment = horizontal_alignment

        print(f"Exported model {model_name} to {output_path}")


if __name__ == "__main__":
    main()
