#!/usr/bin/env python3
#
# Generates an Excel workbook of dbt test failures that can be shared with
# other teams for review and correction.
#
# This script assumes that it is being run in sequence after a call to
# `dbt test --store-failures`, since it depends on two files created by
# that operation (target/run_results.json and target/manifest.json).
# It also requires Python dependencies be installed from
# format_dbt_test_failures_requirements.txt.
#
# Accepts three optional positional arguments:
#
#   1. The local path to a run_results.json file generated by a test run
#     * If not present, defaults to './target/run_results.json'
#   2. The local path to a manifest.json file with the compiled dbt DAG
#     * If not present, defaults to './target/manifest.json'
#   3. The output filepath to which the workbook should be written
#     * If not present, defaults to './qc_test_failures_<date>.xlsx'
#
# Expects one optional environment variable to be set:
#
#  1.AWS_ATHENA_S3_STAGING_DIR: Location in S3 where Athena query results
#    should be written. If missing, defaults to the
#    ccao-athena-results-us-east-1 bucket
#
# Each sheet in the output workbook represents a category of test, e.g.
# "valid_range" or "not_null"; each row in a sheet represents a row in a
# database that failed a test, with enough metadata that a reader can
# figure out what conditions caused the test to fail and investigate the
# root cause.
#
# Example usage:
#
#   AWS_ATHENA_S3_STAGING_DIR=s3://foo-bar-baz/ \
#       python3 format_dbt_test_failures.py \
#       ./target/run_results.json \
#       ./target/manifest.json \
#       ./qc_test_failures.xlsx

import dataclasses
import datetime
import hashlib
import os
import re
import sys
import typing

import openpyxl
import openpyxl.cell
import openpyxl.styles
import openpyxl.styles.colors
import openpyxl.utils
import pyathena
import pyathena.cursor
import simplejson as json

# Tests without a config.meta.category property will be grouped in
# this default category
DEFAULT_TEST_CATEGORY = "miscellaneous"
# Prefix for the URL location of a test in the dbt docs
DOCS_URL_PREFIX = "https://ccao-data.github.io/data-architecture/#!/test"
# The S3 bucket where Athena query results are stored
AWS_ATHENA_S3_STAGING_DIR = os.getenv(
    "AWS_ATHENA_S3_STAGING_DIR", "s3://ccao-athena-results-us-east-1/"
)
# Field names that are used in the output workbook
SOURCE_TABLE_FIELD = "source_table"
DESCRIPTION_FIELD = "description"
TEST_NAME_FIELD = "test_name"
DOCS_URL_FIELD = "docs_url"
TAXYR_FIELD = "taxyr"
PARID_FIELD = "parid"
CARD_FIELD = "card"
TOWNSHIP_FIELD = "township_code"
WHO_FIELD = "who"
WEN_FIELD = "wen"
# Overrides for default display names for dbt tests
CUSTOM_TEST_NAMES = {
    "macro.athena.test_accepted_range": "incorrect_values",
    "macro.dbt_utils.test_accepted_range": "incorrect_values",
    "macro.athena.test_accepted_values": "incorrect_values",
    "macro.athena.test_not_accepted_values": "incorrect_values",
    "macro.dbt_utils.test_not_accepted_values": "incorrect_values",
    "macro.athena.test_unique_combination_of_columns": "duplicate_records",
    "macro.dbt_utils.test_unique_combination_of_columns": "duplicate_records",
    "macro.athena.test_not_null": "missing_values",
    "macro.athena.test_res_class_matches_pardat": "class_mismatch_or_issue",
}
# Directory to store failed test caches
FAILED_TEST_CACHE_DIR = "failed_test_cache"


@dataclasses.dataclass
class FailedTestGroup:
    """Class to store query results for a group of failed dbt tests and provide
    convenience methods for formatting those results for output to a report."""

    # Names of fields that are used for debugging
    _debugging_fieldnames = [TEST_NAME_FIELD, DOCS_URL_FIELD]
    # Names of fields that identify the failing test
    _test_metadata_fieldnames = [
        *[SOURCE_TABLE_FIELD, DESCRIPTION_FIELD],
        *_debugging_fieldnames,
    ]
    # Names of fields that are used for diagnostics
    _diagnostic_fieldnames = [
        TAXYR_FIELD,
        PARID_FIELD,
        CARD_FIELD,
        TOWNSHIP_FIELD,
        WHO_FIELD,
        WEN_FIELD,
    ]
    # The complete set of fixed fields
    _fixed_fieldnames = [
        *_test_metadata_fieldnames,
        *_diagnostic_fieldnames,
    ]

    def __init__(
        self, rows: typing.Optional[typing.List[typing.Dict]] = None
    ) -> None:
        self.raw_rows: typing.List[typing.Dict] = rows or []

    def update(self, new_rows: typing.List[typing.Dict]) -> None:
        """Add a list of new_rows to the rows of failing tests tracked by
        this group. The new_rows list should be formatted like the rows
        returned by a csv.DictReader or a DictCursor, i.e. a list of
        dicts mapping `{column_name: row_value}`."""
        self.raw_rows = [*self.raw_rows, *new_rows]

    @property
    def fieldnames(self) -> typing.List[str]:
        """Get a list of fieldnames that encapsulates all of the fieldnames
        for all of the rows of failing tests tracked by this group."""
        fieldnames = []
        for row in self.raw_rows:
            for column in row.keys():
                if column not in fieldnames:
                    fieldnames.append(column)

        # Remove any fixed fieldnames from the ordered list that are not
        # present in this group
        fixed_field_order = [
            field for field in self._fixed_fieldnames if field in fieldnames
        ]

        # Reorder the fieldnames so that diagnostic fields are presented in the
        # correct order
        for field in reversed(fixed_field_order):
            fieldnames.insert(0, fieldnames.pop(fieldnames.index(field)))

        return fieldnames

    @property
    def rows(self) -> typing.List[typing.List]:
        """Format the rows of failing tests tracked by this group, with
        fieldname data excluded. The combination of this property and the
        `fieldnames` property can be used to write to a csv.Writer or
        to an openpyxl.Workbook sheet for the tests tracked by this group."""
        fieldnames = self.fieldnames
        return [
            [row.get(fieldname) for fieldname in fieldnames]
            for row in self.raw_rows
        ]

    def _filter_for_existing_fieldnames(
        self, possible_fieldnames: typing.List[str]
    ) -> typing.List[str]:
        """Helper function to filter a list of `possible_fieldnames` for
        only those fields that exist in the test group, returning the
        names of the fields (e.g. ["foo", "bar"])."""
        existing_fieldnames = self.fieldnames
        return [
            field
            for field in possible_fieldnames
            if field in existing_fieldnames
        ]

    def _filter_for_existing_field_indexes(
        self, possible_fieldnames: typing.List[str]
    ) -> tuple:
        """Helper function to filter a list of `possible_fieldnames` for
        only those fields that exist in the test group, returning the
        indexes of the fields (e.g. ["A", "B"])."""
        existing_fieldnames = self.fieldnames
        return tuple(
            openpyxl.utils.get_column_letter(
                # openpyxl is 1-indexed while the index() method is 0-indexed
                existing_fieldnames.index(field)
                + 1
            )
            for field in self._filter_for_existing_fieldnames(
                possible_fieldnames
            )
        )

    @property
    def debugging_fieldnames(self) -> typing.List[str]:
        """Get a list of fieldnames (e.g. ["foo", "bar"]) for fields that
        are used for debugging."""
        return self._filter_for_existing_fieldnames(self._debugging_fieldnames)

    @property
    def debugging_field_indexes(self) -> tuple:
        """Get a tuple of field indexes (e.g. ["A", "B"]) for fields that
        are used for debugging."""
        return self._filter_for_existing_field_indexes(
            self._debugging_fieldnames
        )

    @property
    def test_metadata_fieldnames(self) -> typing.List[str]:
        """Get a list of fieldnames (e.g. ["foo", "bar"]) for fields that
        are used for identifying tests."""
        return self._filter_for_existing_fieldnames(
            self._test_metadata_fieldnames
        )

    @property
    def test_metadata_field_indexes(self) -> tuple:
        """Get a tuple of field indexes (e.g. ["A", "B"]) for fields that
        are used for identifying tests."""
        return self._filter_for_existing_field_indexes(
            self._test_metadata_fieldnames
        )

    @property
    def diagnostic_fieldnames(self) -> typing.List[str]:
        """Get a list of fieldnames (e.g. ["foo", "bar"]) for fields that
        are used for diagnostics."""
        return self._filter_for_existing_fieldnames(
            self._diagnostic_fieldnames
        )

    @property
    def diagnostic_field_indexes(self) -> tuple:
        """Get a tuple of field indexes (e.g. ["A", "B"]) for fields that
        are used for diagnostics."""
        return self._filter_for_existing_field_indexes(
            self._diagnostic_fieldnames
        )

    @property
    def fixed_fieldnames(self) -> typing.List[str]:
        """Get a list of fieldnames (e.g. ["foo", "bar"]) for fields that
        are fixed (i.e. whose position is always at the start of the sheet,
        for diagnostic purposes)."""
        return self._filter_for_existing_fieldnames(self._fixed_fieldnames)

    @property
    def fixed_field_indexes(self) -> tuple:
        """Get a list of field indexes (e.g. ["A", "B"]) for fields that
        are fixed (i.e. whose position is always at the start of the sheet,
        for diagnostic purposes)."""
        return self._filter_for_existing_field_indexes(self._fixed_fieldnames)

    @property
    def nonfixed_fieldnames(self) -> typing.List[str]:
        """Get a list of field names (e.g. ["foo", "bar"]) for fields that
        are nonfixed (i.e. whose position comes after the fixed fields in the
        sheet and are thus variable)."""
        fieldnames = self.fieldnames
        fixed_fieldnames = self._fixed_fieldnames
        return [field for field in fieldnames if field not in fixed_fieldnames]

    @property
    def nonfixed_field_indexes(self) -> tuple:
        """Get a list of field indexes (e.g. ["A", "B"]) for fields that
        are nonfixed (i.e. whose position comes after the fixed fields in the
        sheet and are thus variable)."""
        nonfixed_fieldnames = self.nonfixed_fieldnames
        return self._filter_for_existing_field_indexes(nonfixed_fieldnames)


# Type representing a mapping of sheet names to the tests contained therein
FailedTestsByCategory = typing.Dict[str, FailedTestGroup]


def main() -> None:
    """Entrypoint to this script. Parses dbt test failures and writes a
    workbook of test failures to the output path."""
    try:
        run_results_filepath = sys.argv[1]
    except IndexError:
        run_results_filepath = os.path.join("target", "run_results.json")

    try:
        manifest_filepath = sys.argv[2]
    except IndexError:
        manifest_filepath = os.path.join("target", "manifest.json")

    try:
        output_filepath = sys.argv[3]
    except IndexError:
        date_today = datetime.datetime.today().strftime("%Y-%m-%d")
        output_filepath = f"qc_test_failures_{date_today}.xlsx"

    failed_test_cache_path = get_failed_test_cache_path(
        run_results_filepath,
        manifest_filepath,
    )

    if os.path.isfile(failed_test_cache_path):
        print(f"Loading failed tests from cache at {failed_test_cache_path}")
        failed_tests_by_category = get_failed_tests_by_category_from_file(
            failed_test_cache_path
        )
    else:
        print(
            f"Failed test cache not found at {failed_test_cache_path}, "
            "loading failed tests from Athena"
        )
        failed_tests_by_category = get_failed_tests_by_category_from_athena(
            run_results_filepath,
            manifest_filepath,
        )
        print(f"Saving failed tests to the cache at {failed_test_cache_path}")
        save_failed_tests_by_category_to_file(
            failed_tests_by_category,
            failed_test_cache_path,
        )

    print("Generating the output workbook")
    # It's important to use a write-only workbook here because otherwise
    # the metadata required to store cell info about a large number of failing
    # tests can cause the process to run out of memory
    workbook = openpyxl.Workbook(write_only=True)
    for sheet_name, failed_test_group in failed_tests_by_category.items():
        print(f"Adding sheet for {sheet_name}")
        add_sheet_to_workbook(workbook, sheet_name, failed_test_group)
    workbook.save(output_filepath)
    print(f"Output workbook saved to {output_filepath}")


def get_failed_test_cache_path(
    run_results_filepath: str, manifest_filepath: str
) -> str:
    """Return the path to the cache where failed test results are stored.
    The `run_results_filepath` and `manifest_filepath` are used to generated
    a hash key that uniquely defines the cache key for a given test run."""
    with open(run_results_filepath, "rb") as run_results_file:
        run_results_hash = hashlib.md5(run_results_file.read()).hexdigest()

    with open(manifest_filepath, "rb") as manifest_file:
        manifest_hash = hashlib.md5(manifest_file.read()).hexdigest()

    return os.path.join(
        FAILED_TEST_CACHE_DIR,
        f"run_{run_results_hash}_manifest_{manifest_hash}.json",
    )


def get_failed_tests_by_category_from_file(
    file_path: str,
) -> FailedTestsByCategory:
    """Load a FailedTestsByCategory object from a cache located at
    `file_path`."""
    with open(file_path) as cache_file:
        failed_tests_dict = json.load(cache_file, use_decimal=True)
    return {
        category: FailedTestGroup(raw_rows)
        for category, raw_rows in failed_tests_dict.items()
    }


def save_failed_tests_by_category_to_file(
    failed_tests_by_category: FailedTestsByCategory, file_path: str
) -> None:
    """Save a FailedTestsByCategory object to a cache located at
    `file_path`."""
    failed_tests_dict = {
        category: test_group.raw_rows
        for category, test_group in failed_tests_by_category.items()
    }
    os.makedirs(FAILED_TEST_CACHE_DIR, exist_ok=True)
    with open(file_path, "w") as cache_file:
        json.dump(failed_tests_dict, cache_file, use_decimal=True)


def get_failed_tests_by_category_from_athena(
    run_results_filepath: str, manifest_filepath: str
) -> FailedTestsByCategory:
    """Load a FailedTestsByCategory object by querying Athena for failed
    test results generated from a `dbt test --store-failures` call."""
    with open(run_results_filepath) as run_results_fobj:
        run_results = json.load(run_results_fobj)

    with open(manifest_filepath) as manifest_fobj:
        manifest = json.load(manifest_fobj)

    failed_tests_by_category = get_failed_tests_by_category(
        run_results, manifest
    )
    if not failed_tests_by_category:
        raise ValueError(f"{run_results_filepath} contains no failed rows")

    return failed_tests_by_category


def get_failed_tests_by_category(
    run_results: typing.Dict, manifest: typing.Dict
) -> FailedTestsByCategory:
    """Given two artifacts from a `dbt test --store-failures` call (a
    run_results.json file dict and a manifest.json file dict), generates a dict
    where each key is a name of a sheet and each associated value is
    a list of failed tests for that sheet."""
    conn = pyathena.connect(
        s3_staging_dir=AWS_ATHENA_S3_STAGING_DIR,
        region_name="us-east-1",
        cursor_class=pyathena.cursor.DictCursor,
    )
    cursor = conn.cursor()

    failed_tests_by_category: FailedTestsByCategory = {}

    for result in run_results["results"]:
        if result["status"] == "fail":
            unique_id = result["unique_id"]
            # Link to the test's page in the dbt docs, for debugging
            test_docs_url = f"{DOCS_URL_PREFIX}/{unique_id}"

            node = manifest["nodes"].get(unique_id)
            if node is None:
                raise ValueError(
                    f"Missing dbt manifest node with id {unique_id}"
                )

            test_name = node["name"]
            meta = node.get("meta", {})
            category = get_category_from_node(node)
            tablename = get_tablename_from_node(node)
            test_description = meta.get("description")

            # Get the fully-qualified name of the table that stores failures
            # for this test so that we can query it
            test_results_relation_name = node.get("relation_name")
            if test_results_relation_name is None:
                raise ValueError(
                    f"Missing relation_name attribute for test {test_name}. "
                    "Did you run `dbt test` with the --store-failures flag?"
                )

            print(f"Querying failed rows from {test_results_relation_name}")
            # Athena SHOW COLUMNS doesn't allow double quoted tablenames
            relation_name_unquoted = test_results_relation_name.replace(
                '"', "`"
            )
            cursor.execute(f"show columns in {relation_name_unquoted}")
            # SHOW COLUMNS often returns field names with trailing whitespace
            fieldnames = [row["field"].strip() for row in cursor]

            test_results_query = f"select * from {test_results_relation_name}"
            if (
                PARID_FIELD in fieldnames
                and TAXYR_FIELD in fieldnames
                and TOWNSHIP_FIELD not in fieldnames
            ):
                # If parid and taxyr are present, try to retrieve the township
                # code for every row. It's most efficient to do this via a
                # join in the query rather than in a Python lookup since
                # legdat has 43m rows
                test_results_query = f"""
                    select test_results.*, leg.user1 as {TOWNSHIP_FIELD}
                    from {test_results_relation_name} as test_results
                    left join iasworld.legdat as leg
                        on leg.{PARID_FIELD} = test_results.{PARID_FIELD}
                        and leg.{TAXYR_FIELD} = test_results.{TAXYR_FIELD}
                """

            cursor.execute(test_results_query)
            query_results = cursor.fetchall()
            if len(query_results) == 0:
                raise ValueError(
                    f"Test {test_name} has status 'fail' but no failing rows "
                    "in Athena"
                )

            # Add custom fields to query results that we don't expect to be
            # included in the response
            failed_tests = [
                {
                    TEST_NAME_FIELD: test_name,
                    DESCRIPTION_FIELD: test_description,
                    DOCS_URL_FIELD: test_docs_url,
                    SOURCE_TABLE_FIELD: tablename,
                    **row,
                }
                for row in query_results
            ]

            if not failed_tests_by_category.get(category):
                failed_tests_by_category[category] = FailedTestGroup()
            failed_tests_by_category[category].update(failed_tests)

    return failed_tests_by_category


def get_category_from_node(node: typing.Dict) -> str:
    """Given a node representing a dbt test failure, return the category
    that the test should go in."""
    if meta_category := node.get("meta", {}).get("category"):
        return meta_category

    for dependency_macro in node["depends_on"]["macros"]:
        if custom_test_name := CUSTOM_TEST_NAMES.get(dependency_macro):
            return custom_test_name
        # Custom generic tests are always formatted like
        # macro.dbt.test_<generic_name>
        if dependency_macro.startswith("macro.athena.test_"):
            return dependency_macro.split("macro.athena.test_")[-1]
        # dbt_utils generic tests are always formatted like
        # macro.dbt_utils.test_<generic_name>
        if dependency_macro.startswith("macro.dbt_utils.test_"):
            return dependency_macro.split("macro.dbt_utils.test_")[-1]

    return DEFAULT_TEST_CATEGORY


def get_tablename_from_node(node: typing.Dict) -> str:
    """Given a node representing a dbt test failure, return the name of the
    table that the test is testing."""
    if meta_tablename := node.get("meta", {}).get("table_name"):
        # If meta.table_name is set, treat it as an override
        return meta_tablename

    # Search for the model that is implicated in this test via the
    # test_metadata.kwargs.model attribute. Note that it is common to use the
    # elements in the depends_on array for this purpose, but this approach
    # is fraught, since the order of parents for tests with multiple
    # dependencies is not clear and can differ:
    # https://github.com/dbt-labs/dbt-core/issues/6746#issuecomment-1829860236
    test_metadata = node.get("test_metadata", {})
    model_getter_str = test_metadata.get("kwargs", {}).get("model")
    if not model_getter_str:
        raise ValueError(
            "Can't infer tablename: Missing `test_metadata.kwargs.model`"
            f"attribute for test {node['name']}. You may need to add a "
            "`meta.table_name` attribute to the test config to manually "
            "specify the tablename"
        )

    # The test_metadata.kwargs.model attribute is formatted as a Jinja template
    # call to the get_where_subquery macro, so we need to extract the ref or
    # source tablename from that call
    ref_match = re.search(r"ref\('(.+)'\)", model_getter_str)
    if ref_match is not None:
        fq_model_name = ref_match.group(1)
        return fq_model_name.split(".")[-1]

    source_match = re.search(r"source\('iasworld', '(.+)'\)", model_getter_str)
    if source_match is not None:
        return source_match.group(1)

    raise ValueError(
        "Can't infer tablename: Failed to parse model name from "
        f'`test_metadata.kwargs.model` attribute "{model_getter_str}" '
        f" for test \"{node['name']}\". Inspect the dbt manifest file "
        "for more information"
    )


def add_sheet_to_workbook(
    workbook: openpyxl.Workbook,
    sheet_title: str,
    failed_test_group: FailedTestGroup,
) -> None:
    """Add a sheet of failed dbt tests to an openpyxl Workbook. Note that we
    expect the workbook to be initialized with write_only=True."""
    # openpyxl Workbooks are typically created with one untitled active sheet
    # by default, but write-only sheets are an exception to this rule, so we
    # always have to create a new sheet
    sheet = workbook.create_sheet()
    sheet.title = sheet_title

    # Freeze the header row. The syntax for the freeze_panes attribute is
    # undocumented, but it freezes all rows above and all columns to the left
    # of the given cell identifier. Note that freeze operations must be
    # performed before any data is added to a sheet in a write-only workbook
    data_header_idx = 3  # We have three headers; 2 for grouping and 1 for data
    freeze_pane_letter = openpyxl.utils.get_column_letter(
        len(failed_test_group.test_metadata_fieldnames) + 1
    )
    freeze_pane_number = data_header_idx + 1
    sheet.freeze_panes = f"{freeze_pane_letter}{freeze_pane_number}"

    # Hide columns that are intended for debugging only, so that they don't
    # get in the way of non-technical workbook consumers
    for col_idx in failed_test_group.debugging_field_indexes:
        sheet.column_dimensions[col_idx].hidden = True

    # Create groupings for columns with a special group header
    bold_font = openpyxl.styles.Font(bold=True)
    italic_font = openpyxl.styles.Font(italic=True)
    title_row, subtitle_row, header_row, merged_cell_range = [], [], [], []
    column_groups = {
        failed_test_group.test_metadata_field_indexes: {
            "title": "Test description fields",
            "subtitle": "These fields identify a failing test.",
            "fieldnames": failed_test_group.test_metadata_fieldnames,
            "style": "20 % - Accent4",
            "header_style": "Accent4",
        },
        failed_test_group.diagnostic_field_indexes: {
            "title": "Unique identifier fields",
            "subtitle": (
                "These fields identify the row that is failing a test."
            ),
            "fieldnames": failed_test_group.diagnostic_fieldnames,
            "style": "20 % - Accent1",
            "header_style": "Accent1",
        },
        failed_test_group.nonfixed_field_indexes: {
            "title": "Problematic fields",
            "subtitle": (
                "These fields contain values that are causing the test "
                "to fail."
            ),
            "fieldnames": failed_test_group.nonfixed_fieldnames,
            "style": "20 % - Accent2",
            "header_style": "Accent2",
        },
    }
    for col_group_indexes, col_metadata in column_groups.items():
        # Sometimes there are no problematic fields for a given test;
        # if this is the case, skip it
        if not col_group_indexes:
            continue

        # Save merged cell info
        for cell_range in [
            f"{col_group_indexes[0]}1:{col_group_indexes[-1]}1",
            f"{col_group_indexes[0]}2:{col_group_indexes[-1]}2",
        ]:
            merged_cell_range.append(cell_range)

        # Fill out and format grouping header
        title_cell = openpyxl.cell.WriteOnlyCell(
            sheet, value=col_metadata["title"]
        )
        title_cell.style = "Note"
        title_cell.font = bold_font
        title_row.append(title_cell)
        # Flesh out the empty title row cells that will be merged later on
        for _ in range(len(col_group_indexes) - 1):
            title_row.append("")

        subtitle_cell = openpyxl.cell.WriteOnlyCell(
            sheet, value=col_metadata["subtitle"]
        )
        subtitle_cell.style = "Note"
        subtitle_cell.font = italic_font
        subtitle_row.append(subtitle_cell)
        for _ in range(len(col_group_indexes) - 1):
            subtitle_row.append("")

        # Fill out and format the data header
        for fieldname in col_metadata["fieldnames"]:
            header_cell = openpyxl.cell.WriteOnlyCell(sheet, value=fieldname)
            header_cell.style = col_metadata["header_style"]
            header_cell.font = openpyxl.styles.Font(
                bold=True, color=openpyxl.styles.colors.WHITE
            )
            header_row.append(header_cell)

    # Initialize the column widths based on the length of values in
    # the header row
    column_widths = {
        openpyxl.utils.get_column_letter(idx + 1): len(fieldname) + 2
        for idx, fieldname in enumerate(failed_test_group.fieldnames)
    }
    # Iterate the rows to extract data and optionally update the column
    # widths if the length of the cell value exceeds the length of the
    # header value
    data_rows = []
    for row in failed_test_group.rows:
        data_row = []
        # Start enumeration at 1 since openpyxl columns are 1-indexed
        for col_idx, cell in enumerate(row, 1):
            # Convert row values to string so that Excel doesn't apply
            # autoformatting
            cell_str = str(cell) if cell is not None else ""
            cell = openpyxl.cell.WriteOnlyCell(sheet, value=cell_str)

            # Retrieve the cell style from the column groupings if one exists
            cell_style = None
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            for col_group_indexes, col_metadata in column_groups.items():
                if column_letter in col_group_indexes:
                    cell_style = col_metadata["style"]
            if cell_style:
                cell.style = cell_style
            data_row.append(cell)

            # Check if this cell is longer than the longest cell we've seen
            # so far, and adjust the column dimensions accordingly
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            column_widths[column_letter] = max(
                column_widths.get(column_letter, 0), len(cell_str)
            )

        data_rows.append(data_row)

    # Update column widths so that they fit the longest column
    for (
        column_letter,
        column_width,
    ) in column_widths.items():
        # Pad with an extra two characters to account for the fact that
        # non-monospace fonts do not have consistent character widths,
        # and set a hard limit of 75 characters so no one field takes over
        # the viewport of the spreadsheet
        width = min(column_width + 2, 75)
        sheet.column_dimensions[column_letter].width = width

    # Add filters to fixed columns (i.e. columns that appear in every sheet
    # in the same position)
    fixed_field_indexes = failed_test_group.fixed_field_indexes
    sheet_max_row_idx = data_header_idx + len(data_rows)
    min_fixed_idx = f"{fixed_field_indexes[0]}{data_header_idx}"
    max_fixed_idx = f"{fixed_field_indexes[-1]}{sheet_max_row_idx}"
    fixed_field_range = f"{min_fixed_idx}:{max_fixed_idx}"
    sheet.auto_filter.ref = fixed_field_range

    # Add the data to the sheet. This should be one of the last steps in
    # this function, since write-only sheets require all formatting to be
    # set before data is added
    sheet.append(title_row)
    sheet.append(subtitle_row)
    sheet.append(header_row)
    for data_row in data_rows:
        sheet.append(data_row)

    # Merge cells in the grouping headers. This approach is a bit of a hack
    # since merged cells are not fully supported in write-only workbooks, hence
    # why it takes place _after_ rows have been added to the sheet whereas most
    # formatting options for write-only workbooks need to happen _before_
    # data is added. See here for details: https://stackoverflow.com/a/66159254
    for cell_range in merged_cell_range:
        sheet.merged_cells.ranges.add(cell_range)


if __name__ == "__main__":
    main()
